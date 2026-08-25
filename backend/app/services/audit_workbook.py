"""The client deliverable: one workbook, three altitudes, nothing truncated.

WHAT THIS REPLACES. A real 197-page audit rendered an **833-page, 15.4 MB PDF**
listing 8,077 rows. Spec 13.4 names that failure exactly: *"A 300-check audit
that returns a flat severity-sorted list transfers the prioritisation problem to
the client - which is precisely the work an agency is paid to do."*

The same audit through this module is a workbook whose first sheet fits on a
screen and whose last sheet still holds every one of the 8,077 occurrences:

    02_Pillar_Scorecard    6 rows    "where is this site weak, and did we look?"
    10_Findings          461 rows    "Image alt text - 121 pages"        <- fix list
    20_Instances       8,077 rows    "...and here is every one of them"  <- evidence

TWO RULES THAT SHAPE EVERY SHEET.

1. **A score never appears without its coverage.** The same run scored technical
   97.2 having run 25 of 100 technical checks. Every scorecard row carries
   `Checks Ran` and `Checks Applicable` next to the number, and an unmeasured
   dimension prints "not measured", never 0.
2. **The CSV is the complete record.** The XLSX caps instances so it stays
   openable; the CSV never does. When the cap bites, the workbook says so in
   `00_Read_Me` rather than quietly shipping a smaller number.

Reuses the vocabulary already established in ``audit_sheets`` - severity ranks,
priority banding, role mapping, the header styling - so the two deliverables
cannot drift apart.
"""

from __future__ import annotations

import csv
import json
from collections import Counter
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from app.db.database import privileged_connection
from app.services.audit_sheets import (
    _HEADER_FILL,
    _HEADER_FONT,
    _TITLE_FONT,
    _cap,
    _priority_label,
    _priority_score,
    role_for_section,
)

#: The XLSX stops here so the file stays openable on a normal machine; the CSV
#: does not stop at all. Excel's own hard ceiling is 1,048,576 rows.
XLSX_INSTANCE_CAP = 100_000
_FETCH_BATCH = 5_000

WORKBOOK_NAME = "audit-workbook.xlsx"
BUNDLE_NAME = "audit-pack.zip"

#: Dimension -> the section vocabulary `audit_sheets` owns. `pillar` is NOT a
#: valid key for it ("local-seo" is not a section), so keying on pillar silently
#: assigned every local finding to the SEO specialist instead of the local one.
_DIMENSION_TO_SECTION = {
    "onpage": "on-page", "technical": "technical", "offpage": "off-page",
    "local": "local", "geo": "geo", "strategy": "strategy",
}

#: The four checklist files, in operator language rather than filenames.
PILLAR_LABEL: dict[str, str] = {
    "on-page": "On-Page", "technical": "Technical",
    "off-page": "Off-Page", "local-seo": "Local SEO",
}

_DIMENSION_ORDER = ["onpage", "technical", "offpage", "local", "geo", "strategy"]
_SEVERITY_ORDER = ["critical", "major", "minor", "info"]

NOT_MEASURED = "not measured"


@dataclass(slots=True)
class WorkbookResult:
    xlsx: Path | None = None
    bundle: Path | None = None
    csvs: list[Path] = field(default_factory=list)
    findings: int = 0
    instances: int = 0
    instances_in_xlsx: int = 0
    pages: int = 0
    capped: bool = False

    @property
    def files(self) -> list[Path]:
        out = [p for p in (self.xlsx, self.bundle) if p is not None]
        return out + self.csvs


# --------------------------------------------------------------------------- #
# Reading
# --------------------------------------------------------------------------- #

def _fetch_all(cur, sql: str, params: tuple) -> list[dict[str, Any]]:
    cur.execute(sql, params)
    return cur.fetchall()


def _iter_instances(cur, audit_id: str) -> Iterator[dict[str, Any]]:
    """Stream instances in batches - a 500-page site can hold tens of thousands
    and nothing here may materialise them all at once."""
    offset = 0
    while True:
        cur.execute(
            """select i.url, i.instance_kind, i.template_id, i.observed, i.detail,
                      i.evidence, i.severity_override,
                      f.check_id, f.check_name, f.pillar, f.subcategory, f.dimension,
                      f.severity, f.fingerprint
               from public.audit_finding_instances i
               join public.audit_findings f on f.id = i.finding_id
               where i.audit_id = %s
               order by f.check_id, i.url, i.instance_key
               limit %s offset %s""",
            (audit_id, _FETCH_BATCH, offset),
        )
        rows = cur.fetchall()
        if not rows:
            return
        yield from rows
        offset += len(rows)


# --------------------------------------------------------------------------- #
# Cell helpers
# --------------------------------------------------------------------------- #

def _ts(value: Any) -> str:
    """Excel cannot hold a timezone-aware datetime, and Postgres hands us
    ``timestamptz``. Render to a stable ISO-8601 string rather than stripping the
    offset, which would silently move every timestamp for a non-UTC reader."""
    if value is None:
        return ""
    isoformat = getattr(value, "isoformat", None)
    return isoformat(timespec="seconds") if callable(isoformat) else str(value)


def _score_cell(score: Any, checks_ran: int) -> Any:
    """A number, or an explicit statement that we did not look.

    Never 0 for an unmeasured dimension: 0 and "not measured" are opposite
    claims about a client's site and must not share a cell value.
    """
    if score is None or not checks_ran:
        return NOT_MEASURED
    return float(score)


def _coverage_cell(ran: int, applicable: int) -> str:
    return f"{ran} of {applicable}"


def _evidence_text(evidence: Any) -> str:
    if isinstance(evidence, str):
        return _cap(evidence)
    if isinstance(evidence, dict) and evidence:
        return _cap(json.dumps(evidence, ensure_ascii=False, sort_keys=True))
    return ""


# --------------------------------------------------------------------------- #
# Sheet definitions - header, and the row builder for each altitude
# --------------------------------------------------------------------------- #

READ_ME_HEADERS = ("Field", "Value")
EXEC_HEADERS = ("Metric", "Value", "Basis")
PILLAR_HEADERS = (
    "Dimension", "Score", "Coverage", "Checks Ran", "Checks Applicable",
    "Findings", "Instances", "Pages Affected", "Critical", "Major", "Minor", "Info",
    "Not Measured Reason",
)
SUBPOINT_HEADERS = (
    "Pillar", "Subpoint", "Score", "Coverage", "Checks Ran", "Checks Applicable",
    "Findings", "Instances", "Pages Affected",
)
FINDING_HEADERS = (
    "Finding ID", "Check ID", "Check Name", "Pillar", "Subpoint", "Dimension",
    "Owner Agent", "Owner Role", "Automation", "Severity", "Priority",
    "Instances", "Instances Stored", "Pages Affected", "Locus Kind", "Locus Value",
    "Recommendation / Fix", "Evidence", "First Seen", "Last Seen", "Instances Ref",
)
INSTANCE_HEADERS = (
    "Instances Ref", "Check ID", "Check Name", "Pillar", "Subpoint", "Dimension",
    "Severity", "URL", "Template", "Observed", "Detail", "Evidence",
)
PAGE_HEADERS = (
    "URL", "Template", "Page Type", "HTTP", "Indexable", "Depth", "Words",
    "Orphan", "Issues", "Critical", "Major", "Minor", "Info", "Healthy",
)
ROADMAP_HEADERS = (
    "Phase", "Seq", "Title", "Check ID", "Pillar", "Subpoint", "Dimension",
    "Owner Role", "Severity", "Instances", "Pages", "Impact", "Effort Pts",
    "Priority", "Exit Criterion", "Verify With",
)
COVERAGE_HEADERS = (
    "Check ID", "Check Name", "Pillar", "Subpoint", "Dimension", "Owner Agent",
    "Automation", "Severity Default", "Cost Classes", "Data Sources", "Ran?",
    "Skip Reason", "Analyzer Path Resolves",
)


def _pillar_rows(rollups: list[dict]) -> list[list[Any]]:
    by_key = {r["key"]: r for r in rollups if r["level"] == "dimension"}
    rows = []
    for key in _DIMENSION_ORDER:
        r = by_key.get(key)
        if r is None:
            continue
        sev = r.get("severity_counts") or {}
        reason = ""
        if not r["checks_ran"]:
            skips = r.get("skip_reasons") or {}
            reason = ", ".join(f"{k}: {v}" for k, v in sorted(skips.items()))
        rows.append([
            r["label"], _score_cell(r["score"], r["checks_ran"]),
            _coverage_cell(r["checks_ran"], r["checks_applicable"]),
            r["checks_ran"], r["checks_applicable"], r["findings_open"],
            r["instances_open"], r["pages_affected"],
            *[int(sev.get(s, 0)) for s in _SEVERITY_ORDER],
            reason,
        ])
    return rows


def _subpoint_rows(rollups: list[dict]) -> list[list[Any]]:
    rows = []
    for r in rollups:
        if r["level"] != "subpoint":
            continue
        pillar, _, raw = (r["key"] or "").partition("/")
        # `label` is the client-facing subpoint name the engine publishes; `key`
        # is the internal one. A deliverable prints the former.
        sub = r.get("label") or raw
        rows.append([
            PILLAR_LABEL.get(pillar, pillar), sub, _score_cell(r["score"], r["checks_ran"]),
            _coverage_cell(r["checks_ran"], r["checks_applicable"]),
            r["checks_ran"], r["checks_applicable"], r["findings_open"],
            r["instances_open"], r["pages_affected"],
        ])
    # worst measured first, then unmeasured, so the sheet opens on the problems
    rows.sort(key=lambda x: (x[2] == NOT_MEASURED, x[2] if x[2] != NOT_MEASURED else 0))
    return rows


def _finding_rows(findings: list[dict]) -> list[list[Any]]:
    rows = []
    for f in findings:
        score = _priority_score(f["severity"] or "", "fail", None)
        rows.append([
            str(f["id"])[:8], f["check_id"], f["check_name"], f["pillar"],
            f["subcategory"], f["dimension"], f["owner_agent"],
            role_for_section(_DIMENSION_TO_SECTION.get(f["dimension"] or "", "")),
            f["automation"], f["severity"],
            _priority_label(score), f["instance_count"], f["instances_stored"],
            f["pages_affected"], f["locus_kind"], f["locus_value"],
            _cap(f["remediation"] or ""), _evidence_text(f["evidence"]),
            _ts(f["first_seen_at"]), _ts(f["last_seen_at"]), f["fingerprint"],
        ])
    return rows


def _instance_row(i: dict[str, Any]) -> list[Any]:
    return [
        i["fingerprint"], i["check_id"], i["check_name"], i["pillar"],
        i["subcategory"], i["dimension"],
        i["severity_override"] or i["severity"], i["url"], i["template_id"],
        _cap(i["observed"] or ""), _cap(i["detail"] or ""),
        _evidence_text(i["evidence"]),
    ]


def _page_rows(pages: list[dict]) -> list[list[Any]]:
    return [[
        p["url"], p["template_id"], p["page_type"] or "", p["http_status"],
        p["indexable"], p["crawl_depth"], p["word_count"], p["is_orphan"],
        p["issues_total"], p["issues_critical"], p["issues_major"],
        p["issues_minor"], p["issues_info"], p["health_pass"],
    ] for p in pages]


def _roadmap_rows(items: list[dict]) -> list[list[Any]]:
    """The plan, in the order it should be worked.

    Phases are RELATIVE windows, never dates - nothing an audit measures supports
    a calendar claim. The label carries the horizon so the sheet is readable
    without the schema.
    """
    from app.services.audit_roadmap import PHASE_LABEL
    return [[
        PHASE_LABEL.get(i["phase"], i["phase"]), i["sequence"], i["title"],
        i["check_id"], i["pillar"], i["subcategory"], i["dimension"],
        i["owner_role"], i["severity"], i["instance_count"], i["pages_affected"],
        i["impact_score"], i["effort_points"], i["priority"],
        i["exit_criterion"], i["verification_check"],
    ] for i in items]


def _coverage_rows(coverage: dict) -> list[list[Any]]:
    """One row per check in the whole registry - including the ones that did not
    run. This is the sheet that makes "we could not check this" visible, and it
    is also where the cost of each check is legible."""
    checks = coverage.get("checks") or {}
    ran = set(coverage.get("ran") or [])
    skip = {s["check_id"]: s.get("reason", "") for s in coverage.get("skipped") or []}
    rows = []
    for cid in sorted(checks):
        c = checks[cid]
        rows.append([
            cid, c.get("name", ""), c.get("pillar", ""), c.get("subcategory", ""),
            c.get("dimension", ""), c.get("owner_agent", ""), c.get("automation", ""),
            c.get("severity_default", ""), ", ".join(c.get("cost_classes") or []),
            ", ".join(c.get("data_sources") or []),
            "yes" if cid in ran else "no", skip.get(cid, ""),
            "yes" if c.get("analyzer_path_resolves") else "no",
        ])
    return rows


# --------------------------------------------------------------------------- #
# Writing
# --------------------------------------------------------------------------- #

def _write_csv(path: Path, headers: tuple[str, ...], rows) -> int:
    n = 0
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(headers)
        for row in rows:
            w.writerow(row)
            n += 1
    return n



# --------------------------------------------------------------------------- #
# Sheet construction
# --------------------------------------------------------------------------- #
# Every data sheet ships as a real Excel TABLE, not a bare grid.
#
# WHY THAT MATTERS AND IS NOT DECORATION. A bare grid is inert: no filter
# dropdowns, no sort, no banding, no structured reference. The operator's first
# act on receiving one is Ctrl+T - re-doing by hand, per sheet, per audit, work
# the generator can do once. A table also carries its own header row, so freezing
# and filtering survive a re-sort, which a manually frozen pane does not.
#
# The cost is that tables need a NORMAL worksheet; openpyxl's write-only mode
# supports neither tables nor charts. Measured on the real 8,077-instance audit,
# normal mode builds the whole workbook in a few seconds, so the streaming mode
# buys nothing at this scale and costs both features.

_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False,
    showFirstColumn=False, showLastColumn=False,
)

_TITLE_FONT_SHEET = Font(bold=True, size=13, color="6E1423")
_HEADER_ALIGN = Alignment(vertical="center", wrap_text=True)

#: Column widths per header name. Anything unlisted takes the default.
_WIDTHS: dict[str, int] = {
    "URL": 52, "Check Name": 40, "Title": 46, "Recommendation / Fix": 54,
    "Evidence": 40, "Detail": 44, "What was observed": 40, "Observed": 30,
    "Exit Criterion": 46, "Locus Value": 26, "Template": 24, "Subpoint": 26,
    "Skip Reason": 24, "Data Sources": 30, "Field": 30, "Value": 60,
    "Metric": 34, "Basis": 30, "Not Measured Reason": 34, "Coverage": 14,
}
_DEFAULT_WIDTH = 15


def _safe_table_name(sheet: str) -> str:
    """Excel table names cannot start with a digit or contain punctuation, and
    every sheet name here starts with one (`10_Findings`)."""
    return "t_" + "".join(c if c.isalnum() else "_" for c in sheet)


def _write_sheet(
    wb: Workbook, name: str, headers: tuple[str, ...], rows: list[list[Any]],
    *, table: bool = True,
) -> Worksheet:
    """One data sheet: header, rows, widths, frozen header, and a real table."""
    ws = wb.create_sheet(name)
    ws.append(list(headers))
    for row in rows:
        ws.append(row)

    for idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _WIDTHS.get(h, _DEFAULT_WIDTH)
        c = ws.cell(row=1, column=idx)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGN

    ws.freeze_panes = "A2"
    # A table needs at least one body row; a header-only range is invalid and
    # Excel repairs the file by silently discarding it.
    if table and rows:
        ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        t = Table(displayName=_safe_table_name(name), ref=ref)
        t.tableStyleInfo = _TABLE_STYLE
        ws.add_table(t)
    return ws


def _add_score_chart(ws: Worksheet, n_rows: int) -> None:
    """Score by pillar, with coverage beside it.

    The two series are deliberately on ONE chart: a score is not readable without
    the number of checks behind it, and putting them on separate charts is how a
    97 over a quarter of the checklist gets quoted on its own.
    """
    if n_rows < 1:
        return
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Score and coverage by pillar"
    chart.y_axis.title = "Score / % of checks run"
    chart.height, chart.width = 8, 18
    data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=n_rows + 1)
    cov = Reference(ws, min_col=13, max_col=13, min_row=1, max_row=n_rows + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(cov, titles_from_data=True)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    ws.add_chart(chart, f"A{n_rows + 4}")


def _add_severity_chart(ws: Worksheet, anchor_row: int, n: int) -> None:
    """How the issues break down by severity."""
    if n < 1:
        return
    chart = PieChart()
    chart.title = "Issues by severity"
    chart.height, chart.width = 7, 11
    data = Reference(ws, min_col=2, min_row=anchor_row, max_row=anchor_row + n - 1)
    cats = Reference(ws, min_col=1, min_row=anchor_row, max_row=anchor_row + n - 1)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showPercent = True
    ws.add_chart(chart, "E2")


def build(
    *,
    audit_id: str,
    out_dir: str | Path,
    artifact_dir: str | Path | None = None,
    meta: dict[str, Any] | None = None,
) -> WorkbookResult:
    """Build the workbook + CSV pack for one already-ingested audit."""
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    res = WorkbookResult()
    meta = meta or {}

    coverage: dict[str, Any] = {}
    if artifact_dir:
        p = Path(artifact_dir) / "coverage.json"
        if p.is_file():
            try:
                coverage = json.loads(p.read_text(encoding="utf-8"))
            except (ValueError, OSError):
                coverage = {}

    with privileged_connection() as cur:
        rollups = _fetch_all(
            cur,
            "select * from public.audit_rollups where audit_id = %s order by level, key",
            (audit_id,),
        )
        findings = _fetch_all(
            cur,
            # Via this audit's own instances - see audit_findings_repo for why
            # `audit_findings.audit_id` cannot be used for a historical report.
            """select f.* from public.audit_findings f
               where exists (select 1 from public.audit_finding_instances i
                             where i.finding_id = f.id and i.audit_id = %s)
               order by case f.severity when 'critical' then 0 when 'major' then 1
                                        when 'minor' then 2 else 3 end,
                        f.instance_count desc, f.check_id""",
            (audit_id,),
        )
        pages = _fetch_all(
            cur,
            "select * from public.audit_pages where audit_id = %s order by url",
            (audit_id,),
        )

        roadmap_items = _fetch_all(
            cur,
            """select i.* from public.audit_roadmap_items i
               join public.audit_roadmaps m on m.id = i.roadmap_id
               where m.audit_id = %s and m.status = 'active'
               order by case i.phase
                          when 'p0_30d' then 0 when 'p1_90d' then 1
                          when 'p2_180d' then 2 when 'p3_365d' then 3 else 4 end,
                        i.sequence""",
            (audit_id,),
        )

        site = next((r for r in rollups if r["level"] == "site"), {})
        res.findings, res.pages = len(findings), len(pages)

        # ---------------- CSVs first: they are the complete record -----------
        res.csvs.append(out / "findings.csv")
        _write_csv(res.csvs[-1], FINDING_HEADERS, _finding_rows(findings))
        res.csvs.append(out / "pages.csv")
        _write_csv(res.csvs[-1], PAGE_HEADERS, _page_rows(pages))
        res.csvs.append(out / "pillars.csv")
        _write_csv(res.csvs[-1], PILLAR_HEADERS, _pillar_rows(rollups))
        res.csvs.append(out / "subpoints.csv")
        _write_csv(res.csvs[-1], SUBPOINT_HEADERS, _subpoint_rows(rollups))
        if roadmap_items:
            res.csvs.append(out / "roadmap.csv")
            _write_csv(res.csvs[-1], ROADMAP_HEADERS, _roadmap_rows(roadmap_items))
        if coverage:
            res.csvs.append(out / "coverage.csv")
            _write_csv(res.csvs[-1], COVERAGE_HEADERS, _coverage_rows(coverage))

        instances_csv = out / "instances.csv"
        res.instances = _write_csv(
            instances_csv, INSTANCE_HEADERS,
            (_instance_row(i) for i in _iter_instances(cur, audit_id)),
        )
        res.csvs.append(instances_csv)

        # ---------------- XLSX ----------------------------------------------
        # Ordered so the workbook reads front-to-back like a report: what this is,
        # the headline, where the site is weak, what to do, then the evidence.
        wb = Workbook()
        wb.remove(wb.active)

        capped = res.instances > XLSX_INSTANCE_CAP
        res.capped = capped

        # --- 00 Read Me -------------------------------------------------------
        readme_rows: list[list[Any]] = [
            ["Audit", meta.get("url", "")],
            ["Client", meta.get("client_name", "")],
            ["Generated", meta.get("generated_at", "")],
            ["Tier", meta.get("tier", "")],
            ["", ""],
            ["HOW TO READ THIS WORKBOOK", ""],
            ["02_Pillar_Scorecard", "Where the site is weak, and whether we looked"],
            ["03_Subpoint_Scorecard", "The same, one level down"],
            ["05_Roadmap", "What to do first, in relative windows (no dates)"],
            ["10_Issues", "One row per PROBLEM. This is the fix list."],
            ["20_Occurrences", "One row per OCCURRENCE. Join on 'Issue Ref'."],
            ["21_Pages", "Every crawled page, with its issue counts"],
            ["22_Coverage", "Every check in the registry - including what did NOT run"],
            ["", ""],
            ["Checks ran", f"{site.get('checks_ran', 0)} of {site.get('checks_applicable', 0)}"],
            ["Pages crawled", site.get("pages_crawled", 0)],
            ["Issues (distinct problems)", res.findings],
            ["Occurrences", res.instances],
            ["", ""],
            ["IMPORTANT", "A score is only comparable to another score with the SAME basis."],
            ["", "'not measured' means the check did not run - it does NOT mean zero."],
        ]
        if capped:
            readme_rows += [
                ["", ""],
                ["OCCURRENCE CAP APPLIED", f"20_Occurrences shows the first {XLSX_INSTANCE_CAP:,}"],
                ["Complete record", "instances.csv - uncapped"],
            ]
        _write_sheet(wb, "00_Read_Me", READ_ME_HEADERS, readme_rows, table=False)

        # --- 01 Executive summary + severity chart ----------------------------
        basis = f"{site.get('checks_ran', 0)} of {site.get('checks_applicable', 0)} checks"
        exec_rows: list[list[Any]] = [
            ["Overall score", _score_cell(site.get("score"), site.get("checks_ran", 0)), basis],
            ["URL health (critical-free pages)", site.get("url_health_pct"), "denominator: pages crawled"],
            ["Pages crawled", site.get("pages_crawled", 0), ""],
            ["Issues to fix", res.findings, "distinct problems"],
            ["Total occurrences", res.instances, "across all pages"],
            ["Checks run", site.get("checks_ran", 0), f"of {site.get('checks_applicable', 0)} in the registry"],
        ]
        sev_counts = Counter((f["severity"] or "info").lower() for f in findings)
        sev_start = len(exec_rows) + 3
        exec_rows += [["", "", ""], ["Issues by severity", "", ""]]
        for sev in _SEVERITY_ORDER:
            exec_rows.append([sev.capitalize(), sev_counts.get(sev, 0), ""])
        ws = _write_sheet(wb, "01_Executive_Summary", EXEC_HEADERS, exec_rows, table=False)
        ws.cell(row=sev_start, column=1).font = _TITLE_FONT_SHEET
        _add_severity_chart(ws, sev_start + 1, len(_SEVERITY_ORDER))

        # --- 02 Pillar scorecard + chart --------------------------------------
        pillar_rows = _pillar_rows(rollups)
        ws = _write_sheet(wb, "02_Pillar_Scorecard", PILLAR_HEADERS, pillar_rows)
        _add_score_chart(ws, len(pillar_rows))

        # --- 03 / 05 / 10 / 21 ------------------------------------------------
        _write_sheet(wb, "03_Subpoint_Scorecard", SUBPOINT_HEADERS, _subpoint_rows(rollups))
        if roadmap_items:
            _write_sheet(wb, "05_Roadmap", ROADMAP_HEADERS, _roadmap_rows(roadmap_items))
        _write_sheet(wb, "10_Issues", FINDING_HEADERS, _finding_rows(findings))
        _write_sheet(wb, "21_Pages", PAGE_HEADERS, _page_rows(pages))

        # --- 20 Occurrences ---------------------------------------------------
        inst_rows: list[list[Any]] = []
        for inst in _iter_instances(cur, audit_id):
            if len(inst_rows) >= XLSX_INSTANCE_CAP:
                break
            inst_rows.append(_instance_row(inst))
        res.instances_in_xlsx = len(inst_rows)
        _write_sheet(wb, "20_Occurrences", INSTANCE_HEADERS, inst_rows)

        # --- 22 Coverage ------------------------------------------------------
        if coverage:
            _write_sheet(wb, "22_Coverage", COVERAGE_HEADERS, _coverage_rows(coverage))

        res.xlsx = out / WORKBOOK_NAME
        wb.save(res.xlsx)

    # ---------------- bundle: one download, not twenty --------------------
    res.bundle = out / BUNDLE_NAME
    with zipfile.ZipFile(res.bundle, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(res.xlsx, res.xlsx.name)
        for c in res.csvs:
            z.write(c, c.name)
    return res
