"""Data-import worker: streaming, the run-claim, the terminal verdicts, and the
never-re-raise contract.

No DB, no network, no Celery: the privileged store is an in-memory fake and the files are
real files under ``tmp_path`` (the worker's whole job is reading a file, so faking the
filesystem would fake away the thing under test).

The three contracts pinned here, in order of how much they would cost to get wrong:

1. **A redelivered run is a no-op.** ``task_acks_late`` redelivers, and half these targets
   (``backlinks``/``citations``) have NO natural key to conflict on - so a second run
   would duplicate every row. The run-claim is what stops it.
2. **The core never re-raises.** A raise IS the redelivery, so the guard and the
   idempotency are the same property viewed twice.
3. **The file is streamed.** A 200MB export must not be a resident list. Asserted
   structurally (the reader yields) rather than by timing, which would be flaky.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import pytest

from app.config import Settings
from app.modules.data_import.constants import ImportTarget, target_for
from app.modules.data_import.repo import ImportTargetError
from app.modules.data_import.storage import LocalImportStore, iter_rows
from app.modules.data_import.tasks import execute_import

pytestmark = pytest.mark.unit


def _settings(**over: Any) -> Settings:
    return Settings(_env_file=None, app_env="dev", **over)


class FakeStore:
    """In-memory stand-in for the privileged ServiceImportStore.

    ``claim_run`` reproduces the REAL conditional-UPDATE semantics (claim only out of a
    claimable status, flip to 'importing' atomically) - a fake that always handed the run
    over would make every idempotency test below pass vacuously.
    """

    def __init__(self, run: dict[str, Any] | None = None) -> None:
        self.run = run
        self.inserted: list[tuple[str, list[dict[str, Any]]]] = []
        self.progress: list[tuple[int, int, int]] = []
        self.finished: dict[str, Any] | None = None
        self.insert_error: Exception | None = None
        self.batch_sizes: list[int] = []

    def get_run(self, run_id: str) -> dict[str, Any] | None:
        return self.run

    def claim_run(self, run_id: str) -> dict[str, Any] | None:
        if self.run is None or self.run.get("status") not in ("uploaded", "mapping", "validating"):
            return None
        self.run = {**self.run, "status": "importing"}
        return self.run

    def update_progress(self, run_id: str, *, rows_total: int, rows_mapped: int, rows_error: int) -> None:
        self.progress.append((rows_total, rows_mapped, rows_error))

    def finish_run(self, run_id: str, **fields: Any) -> None:
        self.finished = fields

    def insert_rows(self, target: ImportTarget, rows: list[dict[str, Any]]) -> int:
        if self.insert_error is not None:
            raise self.insert_error
        self.batch_sizes.append(len(rows))
        # Copy: the worker clears the batch list in place after flushing.
        self.inserted.append((str(target.table), [dict(r) for r in rows]))
        return len(rows)


def _run_row(**over: Any) -> dict[str, Any]:
    row: dict[str, Any] = {
        "id": "run-1",
        "client_id": "cl-1",
        "client_name": "NorthPeak Dental",
        "filename": "keywords.csv",
        "stored_path": "",
        "source_type": "keywords",
        "status": "mapping",
        "column_map": {"Keyword": "keyword", "Volume": "volume"},
        "detected_columns": ["Keyword", "Volume"],
    }
    row.update(over)
    return row


def _write(tmp_path: Path, store: LocalImportStore, rows: list[list[str]], ext: str = "csv") -> str:
    """Write a real CSV under the store's root and return its key."""
    key = store.new_key(ext)
    with (tmp_path / key).open("w", encoding="utf-8", newline="") as fh:
        csv.writer(fh).writerows(rows)
    return key


@pytest.fixture
def files(tmp_path: Path) -> LocalImportStore:
    return LocalImportStore(tmp_path)


# --------------------------------------------------------------------------- #
# 1. Idempotency - the redelivery contract.
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("terminal", ["imported", "partial", "failed"])
def test_a_redelivered_terminal_run_is_a_no_op(
    tmp_path: Path, files: LocalImportStore, terminal: str
) -> None:
    """THE idempotency test.

    ``acks_late`` redelivers on any raise/timeout. ``backlinks``/``citations`` have no
    unique key to conflict on, so a re-run would duplicate every row the first run wrote.
    A terminal run is unclaimable, so the second delivery writes NOTHING.
    """
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["plumber", "10"]])
    store = FakeStore(_run_row(status=terminal, stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "noop"
    assert store.inserted == []
    assert store.finished is None


def test_a_redelivery_mid_import_is_a_no_op(tmp_path: Path, files: LocalImportStore) -> None:
    """``importing`` is deliberately NOT claimable. Without this a redelivery that lands
    while the first run is still draining would import the same file twice - and the
    append-shaped targets could not dedupe it."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["plumber", "10"]])
    store = FakeStore(_run_row(status="importing", stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "noop"
    assert store.inserted == []


def test_a_second_execution_after_a_successful_one_imports_nothing_more(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """End-to-end idempotency: run the SAME run twice through the same store, exactly as
    a redelivery would, and prove the rows land once."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["plumber", "10"], ["roofer", "20"]])
    store = FakeStore(_run_row(stored_path=key))
    first = execute_import(store, files, _settings(), run_id="run-1")
    assert first["state"] == "imported"
    written = sum(len(rows) for _, rows in store.inserted)

    second = execute_import(store, files, _settings(), run_id="run-1")
    assert second["state"] == "noop"
    assert sum(len(rows) for _, rows in store.inserted) == written  # nothing new


def test_a_missing_run_is_a_no_op_not_a_crash(files: LocalImportStore) -> None:
    store = FakeStore(None)
    result = execute_import(store, files, _settings(), run_id="gone")
    assert result["state"] == "noop"
    assert result["status"] == "missing"


# --------------------------------------------------------------------------- #
# 2. Never re-raises.
# --------------------------------------------------------------------------- #
def test_the_core_never_re_raises_when_the_store_explodes(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """A raise IS a redelivery (acks_late), so the guard and the idempotency are the same
    property. Any unexpected error must become a ``failed`` RESULT, never an exception."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["plumber", "10"]])
    store = FakeStore(_run_row(stored_path=key))
    store.insert_error = ImportTargetError("a programming error")
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"
    assert store.finished is not None
    assert store.finished["status"] == "failed"


def test_the_core_never_re_raises_when_even_the_fail_write_explodes(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """The last line of the guard: if marking the run failed ALSO raises, the task still
    returns - otherwise acks_late would redeliver a run whose rows may already be
    written."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["plumber", "10"]])
    store = FakeStore(_run_row(stored_path=key))
    store.insert_error = RuntimeError("db down")

    def _boom(*a: Any, **k: Any) -> None:
        raise RuntimeError("the fail-write failed too")

    store.finish_run = _boom  # type: ignore[method-assign]
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"


def test_a_batch_the_database_rejects_is_counted_not_fatal(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """A constraint the coercion did not model must not discard the batches that DID
    land: re-raising would redeliver the task and re-insert every successful batch."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["plumber", "10"]])
    store = FakeStore(_run_row(stored_path=key))
    store.insert_error = RuntimeError("23514 check constraint")
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"  # nothing landed at all
    assert result["errors"] == 1


# --------------------------------------------------------------------------- #
# 3. Terminal verdicts: imported | partial | failed.
# --------------------------------------------------------------------------- #
def test_a_clean_file_imports(tmp_path: Path, files: LocalImportStore) -> None:
    key = _write(
        tmp_path, files,
        [["Keyword", "Volume"], ["dental implants", "8100"], ["plumber karachi", "1,200"]],
    )
    store = FakeStore(_run_row(stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result == {"state": "imported", "status": "imported", "rows": 2, "mapped": 2, "errors": 0}
    table, rows = store.inserted[0]
    assert table == "public.keywords"
    assert rows[0] == {
        "keyword": "dental implants", "volume": 8100, "client_id": "cl-1",
        "client_name": "NorthPeak Dental", "source": "import",
    }
    assert rows[1]["volume"] == 1200  # the thousands separator survived the round trip


def test_some_bad_rows_make_the_run_partial_not_failed(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """One bad cell in row 40,000 must not discard the other 39,999."""
    key = _write(
        tmp_path, files,
        [["Keyword", "Volume"], ["good one", "10"], ["bad one", "n/a"], ["good two", "20"]],
    )
    store = FakeStore(_run_row(stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "partial"
    assert (result["rows"], result["mapped"], result["errors"]) == (3, 2, 1)
    assert [r["keyword"] for _, rows in store.inserted for r in rows] == ["good one", "good two"]


def test_a_run_where_nothing_landed_is_failed_not_partial(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """An import that wrote ZERO rows did not partially succeed, whatever the reason -
    and calling it ``partial`` would put a warn tone on a total failure."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["", "10"], ["", "20"]])
    store = FakeStore(_run_row(stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"
    assert result["reason"] == "every row was rejected"


def test_a_header_only_file_is_failed(tmp_path: Path, files: LocalImportStore) -> None:
    key = _write(tmp_path, files, [["Keyword", "Volume"]])
    store = FakeStore(_run_row(stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"
    assert result["reason"] == "the file has no data rows"


def test_an_unmapped_run_fails_before_it_reads_a_row(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """The allow-list gate, re-run at the LAST door before the privileged writer: the run
    row could have been written by an older code path."""
    key = _write(tmp_path, files, [["Keyword"], ["plumber"]])
    store = FakeStore(_run_row(stored_path=key, column_map={}))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"
    assert "invalid column map" in result["reason"]
    assert store.inserted == []


def test_a_persisted_map_naming_a_column_outside_the_allow_list_fails_the_run(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """Defence in depth: even if a hostile map were somehow PERSISTED, the worker
    re-validates and refuses to hand it to the privileged writer."""
    key = _write(tmp_path, files, [["Keyword", "X"], ["plumber", "boom"]])
    store = FakeStore(
        _run_row(stored_path=key, column_map={"Keyword": "keyword", "X": "password_hash"})
    )
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"
    assert "password_hash" in result["reason"]
    assert store.inserted == []


def test_a_missing_file_fails_the_run_cleanly(files: LocalImportStore) -> None:
    store = FakeStore(_run_row(stored_path="0" * 32 + ".csv"))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"
    assert result["reason"] == "the uploaded file is unavailable"


def test_a_traversal_stored_path_fails_the_run_and_reads_nothing(files: LocalImportStore) -> None:
    """The store's guard, seen from the worker: a crafted ``stored_path`` resolves to
    ``None``, so the worker never opens it."""
    store = FakeStore(_run_row(stored_path="../../etc/passwd"))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"
    assert result["reason"] == "the uploaded file is unavailable"


def test_an_unconfigured_store_fails_the_run_rather_than_crashing() -> None:
    store = FakeStore(_run_row(stored_path="x.csv"))
    result = execute_import(store, None, _settings(), run_id="run-1")
    assert result["state"] == "failed"
    assert result["reason"] == "no import store configured"


def test_the_staging_only_custom_type_has_no_downstream_commit(
    tmp_path: Path, files: LocalImportStore
) -> None:
    key = _write(tmp_path, files, [["A"], ["1"]])
    store = FakeStore(_run_row(stored_path=key, source_type="custom", column_map={}))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"
    assert "stage only" in result["reason"]
    assert store.inserted == []


def test_a_rankings_import_without_a_client_fails_cleanly(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """0036's ``tracked_keywords.client_id`` is NOT NULL - a tracked keyword is a standing
    per-client bill. Catching it here beats a 23502 mid-batch."""
    key = _write(tmp_path, files, [["Keyword"], ["plumber"]])
    store = FakeStore(
        _run_row(stored_path=key, source_type="rankings", client_id=None, client_name="",
                 column_map={"Keyword": "keyword"})
    )
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "failed"
    assert "requires a client" in result["reason"]


# --------------------------------------------------------------------------- #
# 4. Streaming + the bounds.
# --------------------------------------------------------------------------- #
def test_the_reader_is_a_generator_that_does_not_slurp(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """Structural proof of streaming: ``iter_rows`` yields, so a 200MB export is never a
    resident list. Asserted by taking ONE row from a 50k-row file - a slurping reader
    would have to build all 50,000 first."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], *[[f"kw{i}", "1"] for i in range(50_000)]])
    rows = iter_rows(tmp_path / key)
    assert next(rows) == ["Keyword", "Volume"]
    assert next(rows) == ["kw0", "1"]
    rows.close()  # type: ignore[union-attr]  # it is a generator: that is the point


def test_a_large_file_is_written_in_bounded_batches(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """Rows are flushed in bounded batches, so memory is O(1) in the file's length rather
    than O(rows) - and progress is streamed as it goes, so a long import is observable
    instead of looking hung."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], *[[f"kw{i}", "1"] for i in range(1_200)]])
    store = FakeStore(_run_row(stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "imported"
    assert result["rows"] == 1_200
    assert store.batch_sizes == [500, 500, 200]  # never one 1,200-row list
    assert len(store.progress) == 3
    assert store.progress[-1] == (1_200, 1_200, 0)


def test_the_row_cap_fails_the_run(tmp_path: Path, files: LocalImportStore) -> None:
    """The worker's runtime bound: without it a crafted 50M-row CSV would hold a Celery
    slot for hours."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], *[[f"kw{i}", "1"] for i in range(20)]])
    store = FakeStore(_run_row(stored_path=key))
    result = execute_import(store, files, _settings(import_max_rows=5), run_id="run-1")
    assert result["state"] == "failed"
    assert "row limit" in result["reason"]


def test_the_error_sample_is_bounded_but_the_counter_is_not(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """A file whose every row is malformed must not write a 200-entry jsonb blob - but
    ``rows_error`` must still report the TRUE total, or a truncated sample would read as
    "only 50 rows were bad"."""
    bad = [[f"kw{i}", "n/a"] for i in range(200)]
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["good", "10"], *bad])
    store = FakeStore(_run_row(stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "partial"
    assert result["errors"] == 200  # every bad row COUNTED
    assert store.finished is not None
    assert len(store.finished["error_sample"]) == 50  # ... but only 50 SAMPLED
    assert store.finished["rows_error"] == 200


def test_the_error_sample_names_the_row_the_column_and_the_reason(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """The sample exists so a human can FIX the file. Row numbers are 1-based INCLUDING
    the header, matching what a spreadsheet shows."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["good", "10"], ["bad", "n/a"]])
    store = FakeStore(_run_row(stored_path=key))
    execute_import(store, files, _settings(), run_id="run-1")
    assert store.finished is not None
    assert store.finished["error_sample"] == [
        {"row": 3, "field": "volume", "value": "n/a", "reason": "not a whole number"}
    ]


def test_blank_separator_lines_are_skipped_not_counted_as_errors(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """Exports routinely carry trailing blank lines; counting them as errors would flag a
    perfectly clean file as ``partial``."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["plumber", "10"], [], ["", ""]])
    store = FakeStore(_run_row(stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result == {"state": "imported", "status": "imported", "rows": 1, "mapped": 1, "errors": 0}


def test_a_short_row_yields_blanks_rather_than_an_index_error(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """Exporters trim trailing empty cells; a zip() that assumed rectangularity would
    crash the whole run on one ragged line."""
    key = _write(tmp_path, files, [["Keyword", "Volume"], ["plumber"]])
    store = FakeStore(_run_row(stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "imported"
    assert store.inserted[0][1][0]["keyword"] == "plumber"
    assert "volume" not in store.inserted[0][1][0]  # left to the DB default


def test_an_excel_bom_does_not_corrupt_the_first_header(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """Excel writes a UTF-8 BOM. Without ``utf-8-sig`` the first header reads as
    ``"\\ufeffKeyword"``, the map misses it, and every row silently fails its required
    field - a whole import lost to three invisible bytes."""
    key = files.new_key("csv")
    (tmp_path / key).write_bytes(b"\xef\xbb\xbfKeyword,Volume\r\nplumber,10\r\n")
    store = FakeStore(_run_row(stored_path=key))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "imported"
    assert store.inserted[0][1][0]["keyword"] == "plumber"


def test_a_real_xlsx_streams_through_the_same_path(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """openpyxl in ``read_only`` mode is what makes the XLSX path a stream too."""
    from openpyxl import Workbook

    key = files.new_key("xlsx")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Keyword", "Volume"])
    ws.append(["dental implants", 8100])
    wb.save(tmp_path / key)

    store = FakeStore(_run_row(stored_path=key, filename="keywords.xlsx"))
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "imported"
    assert store.inserted[0][1][0] == {
        "keyword": "dental implants", "volume": 8100, "client_id": "cl-1",
        "client_name": "NorthPeak Dental", "source": "import",
    }


# --------------------------------------------------------------------------- #
# 5. The derived columns reach the writer.
# --------------------------------------------------------------------------- #
def test_a_citation_import_derives_its_action_from_the_nap_rule(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """The off-page module's own rule, applied by the worker: a spreadsheet cannot set
    ``action`` and therefore cannot contradict the NAP state on the same row."""
    key = _write(
        tmp_path, files,
        [["Directory", "NAP status"], ["Yelp", "missing"], ["Bing Places", "inconsistent"]],
    )
    store = FakeStore(
        _run_row(stored_path=key, source_type="citations",
                 column_map={"Directory": "directory", "NAP status": "nap_status"})
    )
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "imported"
    table, rows = store.inserted[0]
    assert table == "public.citations"
    assert rows[0] == {
        "directory": "Yelp", "nap_status": "missing", "action": "Submit",
        "client_id": "cl-1", "client_name": "NorthPeak Dental",
    }
    assert rows[1]["action"] == "Update"


def test_a_backlinks_import_reduces_the_source_url_to_a_domain(
    tmp_path: Path, files: LocalImportStore
) -> None:
    key = _write(
        tmp_path, files,
        [["Source url", "Page ascore"], ["https://www.example.com/blog/x", "62"]],
    )
    store = FakeStore(
        _run_row(stored_path=key, source_type="backlinks",
                 column_map={"Source url": "ref_domain", "Page ascore": "authority"})
    )
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "imported"
    assert store.inserted[0][1][0] == {
        "ref_domain": "example.com", "authority": 62.0, "client_id": "cl-1",
        "client_name": "NorthPeak Dental",
    }


def test_the_worker_stamps_the_tenant_itself_never_taking_it_from_the_file(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """The writer runs on service_role (BYPASSRLS), so nothing in the DATABASE constrains
    which tenant it writes for - the run row does. A file column can never influence it,
    because ``client_id`` is not in the allow-list at all."""
    key = _write(tmp_path, files, [["Keyword", "client_id"], ["plumber", "cl-ATTACKER"]])
    store = FakeStore(
        _run_row(stored_path=key, column_map={"Keyword": "keyword"},
                 detected_columns=["Keyword", "client_id"])
    )
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "imported"
    written = store.inserted[0][1][0]
    assert written["client_id"] == "cl-1"  # from the RUN
    assert "cl-ATTACKER" not in str(written)


def test_a_search_console_import_links_rows_back_to_their_run(
    tmp_path: Path, files: LocalImportStore
) -> None:
    key = _write(
        tmp_path, files,
        [["Query", "Clicks", "Impressions", "CTR", "Position"],
         ["plumber karachi", "120", "4,000", "3.41%", "8.2"]],
    )
    store = FakeStore(
        _run_row(stored_path=key, source_type="search_console",
                 column_map={"Query": "query", "Clicks": "clicks", "Impressions": "impressions",
                             "CTR": "ctr", "Position": "position"})
    )
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "imported"
    table, rows = store.inserted[0]
    assert table == "public.search_console_rows"
    assert rows[0] == {
        "query": "plumber karachi", "clicks": 120, "impressions": 4000, "ctr": 0.0341,
        "position": 8.2, "client_id": "cl-1", "client_name": "NorthPeak Dental",
        "import_run_id": "run-1",
    }


def test_a_gsc_row_with_neither_query_nor_page_is_rejected_not_ghosted(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """A totals footer has clicks but no query - importing it would create a ghost record
    that inflates every subsequent aggregate."""
    key = _write(
        tmp_path, files,
        [["Query", "Clicks"], ["plumber", "10"], ["", "9999"]],
    )
    store = FakeStore(
        _run_row(stored_path=key, source_type="search_console",
                 column_map={"Query": "query", "Clicks": "clicks"})
    )
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "partial"
    assert (result["mapped"], result["errors"]) == (1, 1)
    assert [r["query"] for _, rows in store.inserted for r in rows] == ["plumber"]


def test_an_agency_global_import_writes_a_null_client(
    tmp_path: Path, files: LocalImportStore
) -> None:
    """0035's bank is client-NULLABLE: an unassigned keyword belongs to no client yet."""
    key = _write(tmp_path, files, [["Keyword"], ["plumber"]])
    store = FakeStore(
        _run_row(stored_path=key, client_id=None, client_name="",
                 column_map={"Keyword": "keyword"})
    )
    result = execute_import(store, files, _settings(), run_id="run-1")
    assert result["state"] == "imported"
    assert store.inserted[0][1][0]["client_id"] is None


def test_the_target_field_target_pairs_are_consistent() -> None:
    """A cheap structural guard: every source_type resolves, and only ``custom`` lacks a
    table (so only ``custom`` can be staging-only)."""
    for source_type in ("keywords", "rankings", "backlinks", "citations", "search_console"):
        target = target_for(source_type)
        assert target is not None and target.table is not None
    custom = target_for("custom")
    assert custom is not None and custom.table is None
