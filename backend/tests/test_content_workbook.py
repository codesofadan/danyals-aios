"""P7: the keyword workbook - one tab per page, and a method trail that survives.

The owner's requirement was a spreadsheet documenting keywords, clusters, strategy and
method "so that we can audit it months later". These tests defend the auditability half,
which is the half that is easy to quietly lose.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from app.services.content_workbook import (
    ESTIMATED_LABEL,
    MAX_SHEET_TITLE,
    cluster_rows,
    data_source,
    master_rows,
    method_rows,
    page_rows,
    render,
    sheet_title,
)


def _term(keyword: str, **kw: Any) -> dict[str, Any]:
    base = {
        "keyword": keyword, "volume": 880, "difficulty": 34.0, "cpc": 12.5,
        "competition": 0.8, "intent": "commercial", "relevance": 0.9,
        "opportunity": 62.0, "cluster_key": "slab-leak", "estimated": False,
    }
    return {**base, **kw}


def _node(keyword: str, **kw: Any) -> dict[str, Any]:
    base = {
        "primary_keyword": keyword, "page_type": "service", "silo": "repairs",
        "intent": "commercial", "target_city": "San Jose", "priority": 1,
        "target_words": 1200, "cluster_key": "slab-leak",
        "secondary_keywords": ["slab leak detection", "under slab pipe repair"],
        "evidence": "3 of the top 10 are directories", "info_gain_thesis": "real prices",
        "published_url": "", "content_job_id": None,
    }
    return {**base, **kw}


def _data(**kw: Any) -> dict[str, Any]:
    base = {
        "engagement": {"client_name": "Delaney Plumbing", "shape": "page_set"},
        "plan": {"provider": "dataforseo", "provider_run_at": "2026-08-25T10:00:00Z",
                 "cost": 1.42, "seed_terms": ["slab leak repair"]},
        "terms": [_term("slab leak repair san jose"), _term("slab leak detection")],
        "nodes": [_node("slab leak repair san jose")],
        "usage": [],
    }
    return {**base, **kw}


# --------------------------------------------------------------------------- #
# Sheet titles - both Excel traps, on real data
# --------------------------------------------------------------------------- #
class TestSheetTitles:
    def test_a_long_keyword_is_trimmed_to_excels_limit(self) -> None:
        long = "emergency slab leak detection and repair in south san jose"
        title = sheet_title(long, set())
        assert len(title) <= MAX_SHEET_TITLE

    @pytest.mark.parametrize("bad", ["roof repair / gutters", "a[b]c", "x:y", "q?", "p*", "s\\t"])
    def test_characters_excel_forbids_are_replaced(self, bad: str) -> None:
        title = sheet_title(bad, set())
        assert not set(title) & set("[]:*?/\\")

    def test_collisions_get_a_prefix_not_a_suffix(self) -> None:
        """A suffix is the first thing the 31-char truncation removes, so it would not
        actually disambiguate the case that produced the collision."""
        long = "emergency slab leak detection and repair south"
        used: set[str] = set()
        first, second = sheet_title(long, used), sheet_title(long, used)
        assert first != second
        assert len(second) <= MAX_SHEET_TITLE
        assert second.startswith("2 ")

    def test_an_empty_name_still_yields_a_legal_title(self) -> None:
        assert sheet_title("", set())


# --------------------------------------------------------------------------- #
# Measured vs estimated - the whole point
# --------------------------------------------------------------------------- #
class TestNumbersCarryTheirProvenance:
    def test_a_derived_figure_is_labelled_estimated(self) -> None:
        """v1 computed `difficulty = log10(totalResults) * 8` and shipped it looking
        exactly like vendor data. A reader must be able to tell them apart."""
        assert data_source(_term("k", estimated=True), "dataforseo") == ESTIMATED_LABEL

    def test_a_bought_figure_names_its_provider(self) -> None:
        assert "dataforseo" in data_source(_term("k"), "dataforseo")

    def test_a_missing_provider_says_so_rather_than_implying_one(self) -> None:
        assert "not recorded" in data_source(_term("k"), "")

    def test_every_master_row_carries_a_data_source(self) -> None:
        rows = master_rows([_term("a"), _term("b", estimated=True)], "dataforseo")
        assert all(row[-1] for row in rows)
        assert ESTIMATED_LABEL in [row[-1] for row in rows]

    def test_cluster_volume_never_mixes_derived_into_a_total(self) -> None:
        """Adding a derived number to a bought one produces a total that is neither -
        and totals are what people quote in meetings."""
        rows = cluster_rows([
            _term("a", volume=100), _term("b", volume=900, estimated=True),
        ])
        assert rows[0][2] == 100, "the estimated 900 must not be in the total"

    def test_an_all_estimated_cluster_reports_no_total_at_all(self) -> None:
        rows = cluster_rows([_term("a", volume=100, estimated=True)])
        assert rows[0][2] == "n/a - all estimated"


class TestMethodAndSources:
    def test_it_records_what_was_bought_from_whom_and_when(self) -> None:
        rows = method_rows(_data()["plan"], _data()["terms"], [])
        first = rows[0]
        assert first[1] == "dataforseo"
        assert first[4] == 1.42
        assert first[5].startswith("2026-08-25")

    def test_derived_terms_get_their_own_row_warning_against_quoting_them(self) -> None:
        terms = [_term("a"), _term("b", estimated=True)]
        rows = method_rows(_data()["plan"], terms, [])
        derived = [r for r in rows if r and "derived" in str(r[0])]
        assert derived, "estimated terms must be called out separately"
        assert "must not be quoted" in derived[0][2]

    def test_writing_spend_is_aggregated_per_stage(self) -> None:
        usage = [
            {"stage": "draft", "model": "claude-sonnet-5", "cost": 0.15},
            {"stage": "draft", "model": "claude-sonnet-5", "cost": 0.20},
            {"stage": "outline", "model": "claude-sonnet-5", "cost": 0.05},
        ]
        rows = method_rows(_data()["plan"], _data()["terms"], usage)
        draft = next(r for r in rows if r and str(r[0]).endswith("draft"))
        assert draft[3] == 2 and round(float(draft[4]), 4) == 0.35

    def test_the_total_includes_both_data_and_writing(self) -> None:
        usage = [{"stage": "draft", "model": "m", "cost": 0.58}]
        rows = method_rows(_data()["plan"], _data()["terms"], usage)
        total = next(r for r in rows if r and r[0] == "TOTAL")
        assert round(float(total[4]), 4) == round(1.42 + 0.58, 4)


class TestPageTabs:
    def test_a_page_tab_states_why_the_page_can_rank(self) -> None:
        """A plan that says WHAT to write without saying WHY is not auditable - months
        later nobody can tell whether the reasoning still holds."""
        rows = page_rows(_node("k"), [_term("k")], "dataforseo")
        labels = [r[0] for r in rows if r]
        assert "Why this page can rank" in labels
        assert "What it adds that the top 10 does not" in labels

    def test_secondary_keywords_are_listed_not_counted(self) -> None:
        rows = page_rows(_node("k"), [], "dataforseo")
        flat = [str(c) for r in rows for c in r]
        assert "slab leak detection" in flat

    def test_an_unpublished_page_says_so_rather_than_showing_blank(self) -> None:
        rows = page_rows(_node("k"), [], "dataforseo")
        url = next(r for r in rows if r and r[0] == "Published URL")
        assert url[1] == "not yet published"


# --------------------------------------------------------------------------- #
# The file itself
# --------------------------------------------------------------------------- #
class TestTheWorkbookOpens:
    def test_it_writes_a_file_openpyxl_can_reopen(self, tmp_path: Path) -> None:
        out = tmp_path / "kw.xlsx"
        result = render(_data(), out)
        assert out.is_file() and result.path == out
        load_workbook(out)

    def test_there_is_one_tab_per_planned_page(self, tmp_path: Path) -> None:
        data = _data(nodes=[_node("slab leak repair"), _node("water heater repair")])
        render(data, tmp_path / "kw.xlsx")
        wb = load_workbook(tmp_path / "kw.xlsx")
        assert "slab leak repair" in wb.sheetnames
        assert "water heater repair" in wb.sheetnames

    def test_the_fixed_tabs_are_all_present(self, tmp_path: Path) -> None:
        render(_data(), tmp_path / "kw.xlsx")
        names = load_workbook(tmp_path / "kw.xlsx").sheetnames
        for expected in ("README", "Method & Sources", "Master Keywords",
                         "Clusters", "Topical Map", "Method Trail"):
            assert expected in names

    def test_two_pages_with_the_same_long_keyword_do_not_collide(self, tmp_path: Path) -> None:
        long = "emergency slab leak detection and repair south san jose"
        render(_data(nodes=[_node(long), _node(long)]), tmp_path / "kw.xlsx")
        names = load_workbook(tmp_path / "kw.xlsx").sheetnames
        assert len(names) == len(set(names)), "duplicate sheet names would lose a page"

    def test_a_keyword_with_a_forbidden_character_still_gets_a_tab(self, tmp_path: Path) -> None:
        render(_data(nodes=[_node("roof repair / gutter cleaning")]), tmp_path / "kw.xlsx")
        load_workbook(tmp_path / "kw.xlsx")

    def test_the_readme_warns_that_estimates_are_not_demand(self, tmp_path: Path) -> None:
        render(_data(terms=[_term("a", estimated=True)]), tmp_path / "kw.xlsx")
        ws = load_workbook(tmp_path / "kw.xlsx")["README"]
        text = " ".join(str(c.value or "") for row in ws.iter_rows() for c in row)
        assert "must not be quoted" in text

    def test_an_empty_engagement_still_produces_a_readable_workbook(self, tmp_path: Path) -> None:
        """A plan with nothing in it is a real state - the operator opens the workbook
        before the research finishes. It must not crash."""
        result = render({"engagement": {}, "plan": {}, "terms": [], "nodes": [],
                         "usage": []}, tmp_path / "kw.xlsx")
        assert result.terms == 0 and result.pages == 0
        load_workbook(tmp_path / "kw.xlsx")

    def test_the_method_trail_shows_dropped_chunks(self, tmp_path: Path) -> None:
        """A doctrine pack that did not fit is a real limit on what the model could
        see. A trail that hides it overstates the doctrine's reach."""
        usage = [{"stage": "draft", "model": "m", "chunk_ids": ["a", "b"],
                  "dropped_chunk_ids": ["c"], "cost": 0.1, "job_id": None}]
        render(_data(usage=usage), tmp_path / "kw.xlsx")
        ws = load_workbook(tmp_path / "kw.xlsx")["Method Trail"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[1][3] == 2 and rows[1][4] == 1
