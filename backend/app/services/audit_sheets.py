"""Deterministic audit remediation sheets: transform a run's ``findings.json``
into role-based fix sheets as ONE ``.xlsx`` workbook (multi-tab) + a set of
single-table ``.csv`` exports.

This is a PURE transform of the audit findings - no AI, no paid provider calls,
no network. Given the same ``findings.json`` it always produces byte-identical
sheets (stable sort, stable ids), exactly like the audit itself.

Deliverables (all written into ``<artifact-root>/<audit_id>/sheets/``):

* ``remediation.xlsx`` - the master workbook. Tabs: ``Summary`` (counts),
  ``Reference`` (every issue, one row), ``Team`` (assignment/tracking master
  list), and one tab per responsible role (a how-to-fix checklist).
* ``summary.csv`` / ``reference.csv`` / ``team.csv`` - the single-table CSV
  exports of those sheets.
* ``role-<slug>.csv`` - one CSV per role (the same rows as its workbook tab).

Design choices (justified):

* ONE combined workbook (tabs) rather than three separate ``.xlsx`` files -
  Excel users expect a single multi-tab file, and per-sheet separation already
  exists as the CSV exports; three redundant workbooks would be worse.
* An ISSUE is a finding whose ``status`` is not a passing/na state (``pass`` /
  ``n_a`` / ``ok`` are dropped - a passing check has nothing to remediate). A
  finding with a missing/unknown status is KEPT (we cannot prove it passed).
* Every issue maps to EXACTLY ONE section and EXACTLY ONE owner role, so the
  role tabs are a strict partition of the reference sheet (nothing dropped,
  nothing double-counted). The mapping table (:data:`SECTION_TO_ROLE`) is
  documented and overridable.

The ``findings.json`` contract is the audit engine's flat JSON ARRAY of finding
objects. The only guaranteed field is ``check_id``; everything else is read
defensively. ``evidence_json`` / ``references_json`` are JSON-ENCODED STRINGS
(a second parse), not nested objects.
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.services.audit_artifacts import LocalArtifactStore
from app.services.evidence_text import humanise_evidence

# --------------------------------------------------------------------------- #
# Roles (the SEO-team owners a finding can be assigned to).
# --------------------------------------------------------------------------- #
ROLE_SEO = "seo_specialist"
ROLE_CONTENT = "content_writer"
ROLE_BLOG = "blog_writer"
ROLE_DEV = "developer"
ROLE_LOCAL = "local_specialist"

# Stable display + iteration order (drives tab order and role-csv order).
ROLE_ORDER: tuple[str, ...] = (ROLE_SEO, ROLE_CONTENT, ROLE_BLOG, ROLE_DEV, ROLE_LOCAL)
ROLE_LABEL: dict[str, str] = {
    ROLE_SEO: "SEO Specialist",
    ROLE_CONTENT: "Content Writer",
    ROLE_BLOG: "Blog Writer",
    ROLE_DEV: "Technical / Developer",
    ROLE_LOCAL: "Local / GBP Specialist",
}
# Short worksheet-tab names (Excel caps tab names at 31 chars; keep them tidy).
ROLE_TAB: dict[str, str] = {
    ROLE_SEO: "SEO Specialist",
    ROLE_CONTENT: "Content Writer",
    ROLE_BLOG: "Blog Writer",
    ROLE_DEV: "Technical",
    ROLE_LOCAL: "Local & GBP",
}
ROLE_CSV: dict[str, str] = {
    ROLE_SEO: "role-seo-specialist.csv",
    ROLE_CONTENT: "role-content-writer.csv",
    ROLE_BLOG: "role-blog-writer.csv",
    ROLE_DEV: "role-developer.csv",
    ROLE_LOCAL: "role-local-specialist.csv",
}

# --------------------------------------------------------------------------- #
# Sections (the 7-way report-contract taxonomy) -> label + role.
# Order = the report-contract order (Strategy, Content, On-page, ...).
# --------------------------------------------------------------------------- #
SECTION_ORDER: tuple[str, ...] = (
    "strategy",
    "content",
    "on-page",
    "technical",
    "off-page",
    "local",
    "geo",
)
SECTION_LABEL: dict[str, str] = {
    "strategy": "Strategy",
    "content": "Content",
    "on-page": "On-Page",
    "technical": "Technical",
    "off-page": "Off-Page",
    "local": "Local SEO",
    "geo": "GEO (AI Search)",
}

# The category->role map. Documented + overridable (pass an override into
# :func:`build_issues`). Rationale:
#   * SEO Specialist owns on-page optimization, off-page/authority, and the
#     derived Strategy narrative (the generalist search owner).
#   * Content Writer owns existing-page content quality (thin/weak copy).
#   * Blog Writer owns GEO / AI-search readiness - citable, direct-answer
#     article content is what earns AI Overview / LLM citations.
#   * Technical / Developer owns crawl/index/infra/schema/security fixes.
#   * Local / GBP Specialist owns Google Business Profile, NAP, reviews.
SECTION_TO_ROLE: dict[str, str] = {
    "strategy": ROLE_SEO,
    "on-page": ROLE_SEO,
    "off-page": ROLE_SEO,
    "content": ROLE_CONTENT,
    "geo": ROLE_BLOG,
    "technical": ROLE_DEV,
    "local": ROLE_LOCAL,
}

# check_id families the report contract re-buckets away from their raw prefix.
_GEO_CHECK_IDS: frozenset[str] = frozenset({"ON-073", "ON-079"})
_OFFPAGE_ON_CHECK_IDS: frozenset[str] = frozenset({"ON-061", "ON-063"})

# --------------------------------------------------------------------------- #
# Severity + status + priority scoring (deterministic).
# --------------------------------------------------------------------------- #
_SEVERITY_RANK: dict[str, int] = {"critical": 0, "major": 1, "minor": 2, "info": 3}
_SEVERITY_LABEL: dict[str, str] = {
    "critical": "Critical",
    "major": "Major",
    "minor": "Minor",
    "info": "Info",
}
_SEVERITY_WEIGHT: dict[str, int] = {"critical": 100, "major": 60, "minor": 30, "info": 10}
_UNKNOWN_SEVERITY_WEIGHT = 40  # unknown/missing severity sits mid-pack

# A finding whose status is one of these is a PASS, not an issue -> excluded.
_NON_ISSUE_STATUSES: frozenset[str] = frozenset({"pass", "passed", "n_a", "na", "ok", "not_applicable"})
_STATUS_WEIGHT: dict[str, int] = {"fail": 20, "warn": 10}
_UNKNOWN_STATUS_WEIGHT = 15

# Effort as a t-shirt size, by the kind of work the section implies.
_EFFORT_BY_SECTION: dict[str, str] = {
    "technical": "L",
    "off-page": "L",
    "content": "M",
    "geo": "M",
    "local": "M",
    "strategy": "M",
    "on-page": "S",
}
_EFFORT_LABEL: dict[str, str] = {"S": "Small", "M": "Medium", "L": "Large"}

_TEXT_CAP = 1500  # cap free-text cells so a pathological evidence blob stays sane

# --------------------------------------------------------------------------- #
# Output filenames (the download allow-list keys off this).
# --------------------------------------------------------------------------- #
XLSX_NAME = "remediation.xlsx"
SUMMARY_CSV = "summary.csv"
REFERENCE_CSV = "reference.csv"
TEAM_CSV = "team.csv"
_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SHEET_FILES: frozenset[str] = frozenset(
    {XLSX_NAME, SUMMARY_CSV, REFERENCE_CSV, TEAM_CSV, *ROLE_CSV.values()}
)


def sheet_media_type(name: str) -> str:
    """Media type for a generated sheet file (xlsx workbook or csv table)."""
    return _XLSX_MEDIA if name.endswith(".xlsx") else "text/csv"


@dataclass(frozen=True)
class SheetMeta:
    """Audit/client context stamped onto the sheets' headers + Summary tab."""

    audit_id: str
    client_name: str
    url: str
    tier: str
    generated_at: str  # ISO-8601


@dataclass(frozen=True)
class Issue:
    """One remediation row - the normalized, sheet-ready view of a finding."""

    row_id: str  # ISS-001, assigned in final priority order
    check_id: str
    section: str  # canonical key (on-page/technical/.../strategy)
    section_label: str
    subcategory: str
    issue: str  # the human check name
    severity: str  # canonical key (critical/major/minor/info/"")
    severity_label: str
    status: str
    url: str
    evidence: str
    fix: str  # the concrete recommendation / how-to-fix
    effort: str  # Small/Medium/Large
    priority_score: int
    priority_label: str  # P1 - Critical ... P4 - Low
    role: str
    role_label: str
    impact_usd: float | None


# --------------------------------------------------------------------------- #
# Classification.
# --------------------------------------------------------------------------- #
def classify_section(finding: dict[str, Any]) -> str:
    """Route a raw finding to one canonical report section (deterministic).

    Uses the ``check_id`` prefix + ``subcategory`` marker first (GEO/off-page
    live under an ``ON-`` prefix but belong to another section), then falls back
    to the raw ``category`` field, then to ``strategy`` for anything unmapped.
    """
    check_id = str(finding.get("check_id") or "").strip().upper()
    subcat = str(finding.get("subcategory") or "").strip().lower()
    category = str(finding.get("category") or "").strip().lower()

    if "geo" in subcat or check_id in _GEO_CHECK_IDS:
        return "geo"
    if check_id.startswith("LOC-") or category in {"local-seo", "local"}:
        return "local"
    if check_id.startswith("OFF-") or check_id in _OFFPAGE_ON_CHECK_IDS or category in {
        "off-page",
        "offpage",
    }:
        return "off-page"
    if check_id.startswith("TECH-") or category == "technical":
        return "technical"
    if category == "content":
        return "content"
    if check_id.startswith("ON-") or category in {"on-page", "onpage"}:
        return "on-page"
    return "strategy"


def role_for_section(section: str, mapping: dict[str, str] | None = None) -> str:
    """The owner role for a section (defaults to SEO Specialist if unmapped)."""
    table = mapping if mapping is not None else SECTION_TO_ROLE
    return table.get(section, ROLE_SEO)


# --------------------------------------------------------------------------- #
# Field extraction helpers.
# --------------------------------------------------------------------------- #
def _is_issue(finding: dict[str, Any]) -> bool:
    status = str(finding.get("status") or "").strip().lower()
    return status not in _NON_ISSUE_STATUSES


def _cap(text: str) -> str:
    text = text.strip()
    return text if len(text) <= _TEXT_CAP else text[: _TEXT_CAP - 1] + "…"


def _evidence_text(finding: dict[str, Any]) -> str:
    """Human evidence string from ``evidence_json`` (a JSON-encoded string).

    Prefers a ``reason`` key; otherwise flattens the object/list. Falls back to
    the ``element`` pointer, then an empty string.
    """
    raw = finding.get("evidence_json")
    data: Any = raw
    if isinstance(raw, str):
        stripped = raw.strip()
        if not stripped:
            data = None
        else:
            try:
                data = json.loads(stripped)
            except (json.JSONDecodeError, ValueError):
                return _cap(stripped)
    if isinstance(data, dict):
        # Rendered through the shared spec: the old `f"{k}: {v}"` flatten put
        # raw keys, nested dicts and internal module paths in front of a client.
        return _cap(humanise_evidence(data))
    if isinstance(data, list):
        return _cap("; ".join(str(x) for x in data))
    if data not in (None, ""):
        return _cap(str(data))
    element = str(finding.get("element") or "").strip()
    return _cap(element)


def _affected_url(finding: dict[str, Any]) -> str:
    url = str(finding.get("url") or "").strip()
    if url:
        return url
    page_id = finding.get("page_id")
    if page_id in (None, ""):
        return "Site-wide"
    return f"Page #{page_id}"


def _fix_text(finding: dict[str, Any], issue_title: str) -> str:
    remediation = str(finding.get("remediation") or "").strip()
    if remediation:
        return _cap(remediation)
    return _cap(f"Review and resolve: {issue_title}")


def _impact_usd(finding: dict[str, Any]) -> float | None:
    raw = finding.get("impact_usd")
    if isinstance(raw, (int, float)):
        return float(raw)
    return None


def _priority_score(severity: str, status: str, impact: float | None) -> int:
    score = _SEVERITY_WEIGHT.get(severity, _UNKNOWN_SEVERITY_WEIGHT)
    score += _STATUS_WEIGHT.get(status, _UNKNOWN_STATUS_WEIGHT)
    if impact and impact > 0:
        score += min(int(impact // 100), 30)
    return score


def _priority_label(score: int) -> str:
    if score >= 100:
        return "P1 - Critical"
    if score >= 70:
        return "P2 - High"
    if score >= 40:
        return "P3 - Medium"
    return "P4 - Low"


def build_issues(
    findings: list[dict[str, Any]], *, role_map: dict[str, str] | None = None
) -> list[Issue]:
    """Normalize raw findings into sorted, id-stamped remediation issues.

    Filters out passing checks, maps each to a section + owner role, scores its
    priority, then sorts by priority (desc), severity, section order, check_id,
    url so the output is deterministic. Row ids (ISS-001..) are assigned AFTER
    the sort, so ISS-001 is always the single most urgent item.
    """
    staged: list[dict[str, Any]] = []
    for finding in findings:
        if not isinstance(finding, dict) or not _is_issue(finding):
            continue
        section = classify_section(finding)
        role = role_for_section(section, role_map)
        severity = str(finding.get("severity") or "").strip().lower()
        status = str(finding.get("status") or "").strip().lower()
        impact = _impact_usd(finding)
        title = str(finding.get("check_name") or finding.get("check_id") or "Issue").strip()
        score = _priority_score(severity, status, impact)
        staged.append(
            {
                "check_id": str(finding.get("check_id") or "").strip(),
                "section": section,
                "subcategory": str(finding.get("subcategory") or "").strip(),
                "issue": title,
                "severity": severity,
                "status": status,
                "url": _affected_url(finding),
                "evidence": _evidence_text(finding),
                "fix": _fix_text(finding, title),
                "effort": _EFFORT_LABEL[_EFFORT_BY_SECTION.get(section, "M")],
                "priority_score": score,
                "role": role,
                "impact_usd": impact,
            }
        )

    section_index = {name: i for i, name in enumerate(SECTION_ORDER)}
    staged.sort(
        key=lambda s: (
            -int(s["priority_score"]),
            _SEVERITY_RANK.get(s["severity"], 4),
            section_index.get(s["section"], len(SECTION_ORDER)),
            s["check_id"],
            s["url"],
            s["issue"],
        )
    )

    issues: list[Issue] = []
    for i, s in enumerate(staged, start=1):
        severity = str(s["severity"])
        section = str(s["section"])
        role = str(s["role"])
        issues.append(
            Issue(
                row_id=f"ISS-{i:03d}",
                check_id=str(s["check_id"]),
                section=section,
                section_label=SECTION_LABEL.get(section, section.title()),
                subcategory=str(s["subcategory"]),
                issue=str(s["issue"]),
                severity=severity,
                severity_label=_SEVERITY_LABEL.get(severity, severity.title() or "-"),
                status=str(s["status"]),
                url=str(s["url"]),
                evidence=str(s["evidence"]),
                fix=str(s["fix"]),
                effort=str(s["effort"]),
                priority_score=int(s["priority_score"]),
                priority_label=_priority_label(int(s["priority_score"])),
                role=role,
                role_label=ROLE_LABEL.get(role, role),
                impact_usd=s["impact_usd"],
            )
        )
    return issues


# --------------------------------------------------------------------------- #
# Counts / summary.
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class SheetCounts:
    total_findings: int
    total_issues: int
    by_severity: dict[str, int]
    by_section: dict[str, int]
    by_role: dict[str, int]
    by_priority: dict[str, int]


_PRIORITY_BANDS: tuple[str, ...] = ("P1 - Critical", "P2 - High", "P3 - Medium", "P4 - Low")


def compute_counts(issues: list[Issue], total_findings: int) -> SheetCounts:
    by_severity = dict.fromkeys(_SEVERITY_LABEL, 0)
    by_section = dict.fromkeys(SECTION_ORDER, 0)
    by_role = dict.fromkeys(ROLE_ORDER, 0)
    by_priority = dict.fromkeys(_PRIORITY_BANDS, 0)
    for it in issues:
        by_severity[it.severity] = by_severity.get(it.severity, 0) + 1
        by_section[it.section] = by_section.get(it.section, 0) + 1
        by_role[it.role] = by_role.get(it.role, 0) + 1
        by_priority[it.priority_label] = by_priority.get(it.priority_label, 0) + 1
    return SheetCounts(
        total_findings=total_findings,
        total_issues=len(issues),
        by_severity=by_severity,
        by_section=by_section,
        by_role=by_role,
        by_priority=by_priority,
    )


# --------------------------------------------------------------------------- #
# Column definitions (shared by the xlsx tabs and the csv exports).
# --------------------------------------------------------------------------- #
REFERENCE_HEADERS: tuple[str, ...] = (
    "ID",
    "Check ID",
    "Category",
    "Subcategory",
    "Issue",
    "Severity",
    "Status",
    "Affected URL / Page",
    "Evidence",
    "Recommendation / Fix",
    "Effort",
    "Priority",
    "Owner Role",
    "Est. $ Impact / mo",
)
TEAM_HEADERS: tuple[str, ...] = (
    "ID",
    "Category",
    "Issue",
    "Severity",
    "Priority",
    "Owner Role",
    "Effort",
    "Status",
    "Assignee",
    "Target Date",
    "Notes",
)
ROLE_HEADERS: tuple[str, ...] = (
    "Done",
    "ID",
    "Priority",
    "Severity",
    "Category",
    "Issue",
    "Affected URL / Page",
    "How to Fix",
    "Evidence",
    "Effort",
)

_TRACKING_STATES = "Not started,In progress,Blocked,In review,Done"


def _impact_cell(value: float | None) -> str:
    return f"${value:,.0f}" if value else ""


def _reference_row(it: Issue) -> list[Any]:
    return [
        it.row_id,
        it.check_id,
        it.section_label,
        it.subcategory,
        it.issue,
        it.severity_label,
        it.status,
        it.url,
        it.evidence,
        it.fix,
        it.effort,
        it.priority_label,
        it.role_label,
        _impact_cell(it.impact_usd),
    ]


def _team_row(it: Issue) -> list[Any]:
    return [
        it.row_id,
        it.section_label,
        it.issue,
        it.severity_label,
        it.priority_label,
        it.role_label,
        it.effort,
        "Not started",  # tracking column the lead fills in
        "",  # Assignee
        "",  # Target Date
        "",  # Notes
    ]


def _role_row(it: Issue) -> list[Any]:
    return [
        "",  # Done checkbox column
        it.row_id,
        it.priority_label,
        it.severity_label,
        it.section_label,
        it.issue,
        it.url,
        it.fix,
        it.evidence,
        it.effort,
    ]


# --------------------------------------------------------------------------- #
# CSV writers.
# --------------------------------------------------------------------------- #
def _write_csv(path: Path, headers: tuple[str, ...], rows: list[list[Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        writer.writerows(rows)


def write_reference_csv(issues: list[Issue], path: Path) -> None:
    _write_csv(path, REFERENCE_HEADERS, [_reference_row(it) for it in issues])


def write_team_csv(issues: list[Issue], path: Path) -> None:
    _write_csv(path, TEAM_HEADERS, [_team_row(it) for it in issues])


def write_role_csv(role: str, issues: list[Issue], path: Path) -> None:
    rows = [_role_row(it) for it in issues if it.role == role]
    _write_csv(path, ROLE_HEADERS, rows)


def write_summary_csv(counts: SheetCounts, meta: SheetMeta, path: Path) -> None:
    rows: list[list[Any]] = [
        ["Section", "Metric", "Value"],
        ["Audit", "Audit ID", meta.audit_id],
        ["Audit", "Client", meta.client_name],
        ["Audit", "URL", meta.url],
        ["Audit", "Tier", meta.tier],
        ["Audit", "Generated", meta.generated_at],
        ["Totals", "Findings analyzed", counts.total_findings],
        ["Totals", "Remediation issues", counts.total_issues],
    ]
    for key in _SEVERITY_LABEL:
        rows.append(["By severity", _SEVERITY_LABEL[key], counts.by_severity.get(key, 0)])
    for key in SECTION_ORDER:
        rows.append(["By category", SECTION_LABEL[key], counts.by_section.get(key, 0)])
    for key in ROLE_ORDER:
        rows.append(["By owner role", ROLE_LABEL[key], counts.by_role.get(key, 0)])
    for band in _PRIORITY_BANDS:
        rows.append(["By priority", band, counts.by_priority.get(band, 0)])
    with path.open("w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(rows)


# --------------------------------------------------------------------------- #
# XLSX workbook.
# --------------------------------------------------------------------------- #
_HEADER_FILL = PatternFill("solid", fgColor="6E1423")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color="6E1423")
_SEVERITY_FILL: dict[str, PatternFill] = {
    "critical": PatternFill("solid", fgColor="F4CCCC"),
    "major": PatternFill("solid", fgColor="FCE5CD"),
    "minor": PatternFill("solid", fgColor="FFF2CC"),
    "info": PatternFill("solid", fgColor="EFEFEF"),
}
_WRAP = Alignment(vertical="top", wrap_text=True)

# Per-header column widths (best-effort; falls back to a default).
_COL_WIDTH: dict[str, int] = {
    "ID": 9,
    "Check ID": 10,
    "Category": 14,
    "Subcategory": 14,
    "Issue": 40,
    "Severity": 11,
    "Status": 8,
    "Affected URL / Page": 34,
    "Evidence": 46,
    "Recommendation / Fix": 52,
    "How to Fix": 52,
    "Effort": 9,
    "Priority": 14,
    "Owner Role": 20,
    "Est. $ Impact / mo": 15,
    "Done": 7,
    "Assignee": 16,
    "Target Date": 13,
    "Notes": 24,
}
_DEFAULT_WIDTH = 16
_WIDE_TEXT_HEADERS = frozenset({"Issue", "Evidence", "Recommendation / Fix", "How to Fix"})


def _write_data_sheet(
    ws: Any, title: str, headers: tuple[str, ...], rows: list[list[Any]], severity_col: int | None
) -> None:
    """Write a titled, styled table: row 1 = title, row 2 = headers, row 3+ data."""
    ncols = len(headers)
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = _COL_WIDTH.get(header, _DEFAULT_WIDTH)
    for r, row in enumerate(rows, start=3):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            if headers[col - 1] in _WIDE_TEXT_HEADERS:
                cell.alignment = _WRAP
        if severity_col is not None:
            sev = str(row[severity_col]).strip().lower()
            fill = _SEVERITY_FILL.get(sev)
            if fill is not None:
                ws.cell(row=r, column=severity_col + 1).fill = fill
    ws.freeze_panes = "A3"
    last_row = max(2, len(rows) + 2)
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{last_row}"


def _write_summary_tab(ws: Any, counts: SheetCounts, meta: SheetMeta) -> None:
    ws.cell(row=1, column=1, value="Audit Remediation - Summary").font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    meta_rows = [
        ("Client", meta.client_name),
        ("URL", meta.url),
        ("Audit ID", meta.audit_id),
        ("Tier", meta.tier),
        ("Generated", meta.generated_at),
        ("Findings analyzed", counts.total_findings),
        ("Remediation issues", counts.total_issues),
    ]
    r = 3
    for label, value in meta_rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        r += 1

    def _block(title: str, pairs: list[tuple[str, int]], start: int) -> int:
        cell = ws.cell(row=start, column=1, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.cell(row=start, column=2, value="Count").font = _HEADER_FONT
        ws.cell(row=start, column=2).fill = _HEADER_FILL
        row = start + 1
        for name, count in pairs:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=count)
            row += 1
        return row + 1

    r += 1
    r = _block("By severity", [(_SEVERITY_LABEL[k], counts.by_severity.get(k, 0)) for k in _SEVERITY_LABEL], r)
    r = _block("By category", [(SECTION_LABEL[k], counts.by_section.get(k, 0)) for k in SECTION_ORDER], r)
    r = _block("By owner role", [(ROLE_LABEL[k], counts.by_role.get(k, 0)) for k in ROLE_ORDER], r)
    _block("By priority", [(b, counts.by_priority.get(b, 0)) for b in _PRIORITY_BANDS], r)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40


def build_workbook(issues: list[Issue], meta: SheetMeta, total_findings: int) -> Any:
    """Assemble the combined ``remediation.xlsx`` workbook (returns a Workbook)."""
    counts = compute_counts(issues, total_findings)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    _write_summary_tab(summary, counts, meta)

    ref = wb.create_sheet("Reference")
    _write_data_sheet(
        ref,
        f"Reference - All Issues ({len(issues)})",
        REFERENCE_HEADERS,
        [_reference_row(it) for it in issues],
        severity_col=REFERENCE_HEADERS.index("Severity"),
    )

    team = wb.create_sheet("Team")
    team_rows = [_team_row(it) for it in issues]
    _write_data_sheet(
        team,
        f"Team Assignment & Tracking ({len(issues)})",
        TEAM_HEADERS,
        team_rows,
        severity_col=TEAM_HEADERS.index("Severity"),
    )
    if team_rows:
        dv = DataValidation(type="list", formula1=f'"{_TRACKING_STATES}"', allow_blank=True)
        team.add_data_validation(dv)
        status_letter = get_column_letter(TEAM_HEADERS.index("Status") + 1)
        dv.add(f"{status_letter}3:{status_letter}{len(team_rows) + 2}")

    for role in ROLE_ORDER:
        role_rows = [_role_row(it) for it in issues if it.role == role]
        ws = wb.create_sheet(ROLE_TAB[role])
        _write_data_sheet(
            ws,
            f"{ROLE_LABEL[role]} - Fix Checklist ({len(role_rows)})",
            ROLE_HEADERS,
            role_rows,
            severity_col=ROLE_HEADERS.index("Severity"),
        )
    return wb


# --------------------------------------------------------------------------- #
# Generation + storage.
# --------------------------------------------------------------------------- #
def generate_sheets(findings: list[dict[str, Any]], meta: SheetMeta, out_dir: Path) -> list[str]:
    """Build every sheet from parsed findings; return the filenames written."""
    issues = build_issues(findings)
    counts = compute_counts(issues, len(findings))
    out_dir.mkdir(parents=True, exist_ok=True)

    build_workbook(issues, meta, len(findings)).save(out_dir / XLSX_NAME)
    write_summary_csv(counts, meta, out_dir / SUMMARY_CSV)
    write_reference_csv(issues, out_dir / REFERENCE_CSV)
    write_team_csv(issues, out_dir / TEAM_CSV)
    written = [XLSX_NAME, SUMMARY_CSV, REFERENCE_CSV, TEAM_CSV]
    for role in ROLE_ORDER:
        name = ROLE_CSV[role]
        write_role_csv(role, issues, out_dir / name)
        written.append(name)
    return written


def load_findings(src: str | Path | None) -> list[dict[str, Any]] | None:
    """Parse a ``findings.json`` file into a list of dicts, or ``None``.

    Degrade-safe: a missing file, unreadable file, invalid JSON, or a top-level
    value that is not a JSON array all return ``None`` (skip sheet generation) -
    never an exception, so the audit run is never failed by a bad findings file.
    """
    if not src:
        return None
    path = Path(src)
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None
    if not isinstance(data, list):
        return None
    return [d for d in data if isinstance(d, dict)]


def store_audit_sheets(
    store: LocalArtifactStore, audit_id: str, findings_src: str | Path | None, meta: SheetMeta
) -> list[str]:
    """Generate + persist the sheets under ``<root>/<audit_id>/sheets/``.

    Returns the filenames written, or ``[]`` when findings are missing/malformed
    (degrade-safe). Any exception here is the caller's to swallow - a sheet build
    must NEVER fail the audit job.
    """
    findings = load_findings(findings_src)
    if findings is None:
        return []
    return generate_sheets(findings, meta, store.sheets_dir(audit_id))
