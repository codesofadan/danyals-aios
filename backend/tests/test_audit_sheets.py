"""Gate for the audit remediation sheets: the deterministic builder + the guarded
staff/portal download routes.

The builder is proven against a representative ``findings.json`` fixture (real
engine schema): every issue appears once in the reference sheet, the role tabs
strictly partition the issues (none dropped, each mapped to a role), the team
sheet lists all with a role + priority, the xlsx opens with the expected tabs +
cells, and csv row counts match. The routes mirror the report.pdf/findings.json
auth + ownership guards.
"""

from __future__ import annotations

import json
from collections.abc import Callable
from pathlib import Path
from typing import Any

import httpx
import pytest
from fastapi import FastAPI
from openpyxl import load_workbook

from app.core.auth import CurrentUser, get_current_user
from app.db.audits_repo import get_audits_repo
from app.db.portal_repo import get_portal_repo
from app.routers.audits import get_artifact_store
from app.services.audit_artifacts import LocalArtifactStore
from app.services.audit_sheets import (
    REFERENCE_HEADERS,
    ROLE_BLOG,
    ROLE_CONTENT,
    ROLE_CSV,
    ROLE_DEV,
    ROLE_LOCAL,
    ROLE_ORDER,
    ROLE_SEO,
    ROLE_TAB,
    SHEET_FILES,
    SUMMARY_CSV,
    TEAM_HEADERS,
    XLSX_NAME,
    SheetMeta,
    build_issues,
    classify_section,
    generate_sheets,
    load_findings,
    role_for_section,
    store_audit_sheets,
)

pytestmark = pytest.mark.unit

_FIXTURE = Path(__file__).parent / "fixtures" / "findings_sample.json"

# The fixture has 13 findings; 2 are non-issues (TECH-099 pass, LOC-050 n_a).
_EXPECTED_ISSUE_COUNT = 11
_EXPECTED_ROLE_COUNTS = {
    ROLE_SEO: 4,  # ON-005, OFF-007, ON-061, STRAT-002
    ROLE_CONTENT: 1,  # CNT-014
    ROLE_BLOG: 1,  # ON-073 (geo-ai)
    ROLE_DEV: 3,  # TECH-021, TECH-045, TECH-030
    ROLE_LOCAL: 2,  # LOC-001, LOC-070
}


def _findings() -> list[dict[str, Any]]:
    data = load_findings(_FIXTURE)
    assert data is not None
    return data


def _meta() -> SheetMeta:
    return SheetMeta(
        audit_id="aud-1",
        client_name="Acme Co",
        url="https://acme.example/",
        tier="Free",
        generated_at="2026-08-12T10:00:00+00:00",
    )


def _read_csv(path: Path) -> list[list[str]]:
    import csv

    with path.open(newline="", encoding="utf-8") as fh:
        return list(csv.reader(fh))


# --- classification / role mapping --------------------------------------------
def test_classify_section_routes_by_prefix_and_subcategory() -> None:
    assert classify_section({"check_id": "TECH-021", "category": "technical"}) == "technical"
    assert classify_section({"check_id": "LOC-001", "category": "local-seo"}) == "local"
    assert classify_section({"check_id": "OFF-007", "category": "off-page"}) == "off-page"
    assert classify_section({"check_id": "ON-005", "category": "on-page"}) == "on-page"
    # GEO is not a category - it's an ON- check with a geo-ai subcategory.
    assert classify_section({"check_id": "ON-073", "subcategory": "geo-ai"}) == "geo"
    # Content is a first-class category.
    assert classify_section({"check_id": "CNT-014", "category": "content"}) == "content"
    # A linking ON- check that the report contract re-buckets to off-page.
    assert classify_section({"check_id": "ON-061", "category": "on-page"}) == "off-page"
    # Anything unmapped rolls up to strategy.
    assert classify_section({"check_id": "STRAT-002", "category": "strategy"}) == "strategy"
    assert classify_section({"check_id": "???"}) == "strategy"


def test_role_for_section_maps_every_section() -> None:
    assert role_for_section("technical") == ROLE_DEV
    assert role_for_section("local") == ROLE_LOCAL
    assert role_for_section("content") == ROLE_CONTENT
    assert role_for_section("geo") == ROLE_BLOG
    assert role_for_section("on-page") == ROLE_SEO
    assert role_for_section("off-page") == ROLE_SEO
    assert role_for_section("strategy") == ROLE_SEO


# --- build_issues -------------------------------------------------------------
def test_build_issues_excludes_passing_checks() -> None:
    issues = build_issues(_findings())
    assert len(issues) == _EXPECTED_ISSUE_COUNT
    check_ids = {it.check_id for it in issues}
    # The pass + n_a findings are not remediation items.
    assert "TECH-099" not in check_ids
    assert "LOC-050" not in check_ids
    # A finding with a MISSING status is kept (we cannot prove it passed).
    assert "TECH-030" in check_ids


def test_build_issues_are_priority_sorted_with_stable_ids() -> None:
    issues = build_issues(_findings())
    # ISS-001 is the single most urgent item (a critical).
    assert issues[0].row_id == "ISS-001"
    assert issues[0].severity == "critical"
    # Row ids are sequential + priority scores are monotonically non-increasing.
    assert [it.row_id for it in issues] == [f"ISS-{i:03d}" for i in range(1, len(issues) + 1)]
    scores = [it.priority_score for it in issues]
    assert scores == sorted(scores, reverse=True)
    # Every issue carries a concrete fix + a P-band.
    assert all(it.fix for it in issues)
    assert all(it.priority_label.startswith("P") for it in issues)


def test_evidence_and_url_extraction() -> None:
    by_check = {it.check_id: it for it in build_issues(_findings())}
    # evidence_json is a JSON-encoded string; we surface the 'reason'.
    assert "plain http" in by_check["TECH-021"].evidence
    # A malformed evidence_json degrades to the raw string, not a crash.
    assert by_check["TECH-045"].evidence == "malformed-not-json"
    # A finding with no url + null page_id is Site-wide.
    assert by_check["LOC-070"].url == "Site-wide"
    # $ impact is carried through.
    assert by_check["TECH-021"].impact_usd == 1200.0


# --- role partition -----------------------------------------------------------
def test_roles_strictly_partition_all_issues() -> None:
    issues = build_issues(_findings())
    buckets: dict[str, list[str]] = {role: [] for role in ROLE_ORDER}
    for it in issues:
        assert it.role in ROLE_ORDER  # every issue has a known role
        buckets[it.role].append(it.row_id)

    # Disjoint + covering: the union of the role buckets == all issues, no dupes.
    all_ids = [rid for ids in buckets.values() for rid in ids]
    assert len(all_ids) == len(issues)
    assert set(all_ids) == {it.row_id for it in issues}
    assert len(set(all_ids)) == len(all_ids)  # nothing double-counted
    assert {role: len(ids) for role, ids in buckets.items()} == _EXPECTED_ROLE_COUNTS


# --- xlsx workbook ------------------------------------------------------------
def test_xlsx_opens_with_expected_tabs_and_cells(tmp_path: Path) -> None:
    generate_sheets(_findings(), _meta(), tmp_path)
    wb = load_workbook(tmp_path / XLSX_NAME)
    expected_tabs = ["Summary", "Reference", "Team", *[ROLE_TAB[r] for r in ROLE_ORDER]]
    assert wb.sheetnames == expected_tabs

    ref = wb["Reference"]
    # Row 1 = title, row 2 = headers, row 3+ = data.
    assert list(REFERENCE_HEADERS) == [ref.cell(row=2, column=c).value for c in range(1, len(REFERENCE_HEADERS) + 1)]
    assert ref.cell(row=3, column=1).value == "ISS-001"
    # One data row per issue (max_row = 2 header rows + N issues).
    assert ref.max_row == _EXPECTED_ISSUE_COUNT + 2

    team = wb["Team"]
    assert list(TEAM_HEADERS) == [team.cell(row=2, column=c).value for c in range(1, len(TEAM_HEADERS) + 1)]
    # The tracking Status column defaults to "Not started".
    status_col = TEAM_HEADERS.index("Status") + 1
    assert team.cell(row=3, column=status_col).value == "Not started"

    # A role tab holds exactly its share of issues.
    dev = wb[ROLE_TAB[ROLE_DEV]]
    assert dev.max_row == _EXPECTED_ROLE_COUNTS[ROLE_DEV] + 2


# --- csv exports --------------------------------------------------------------
def test_csv_row_counts_match(tmp_path: Path) -> None:
    written = generate_sheets(_findings(), _meta(), tmp_path)
    # Exactly the allow-listed files, nothing more/less.
    assert set(written) == set(SHEET_FILES)

    ref_rows = _read_csv(tmp_path / "reference.csv")
    assert ref_rows[0] == list(REFERENCE_HEADERS)  # header
    assert len(ref_rows) - 1 == _EXPECTED_ISSUE_COUNT  # one data row per issue

    # Per-role csv counts sum to the whole, matching the partition.
    role_total = 0
    for role in ROLE_ORDER:
        rows = _read_csv(tmp_path / ROLE_CSV[role])
        data_rows = len(rows) - 1
        assert data_rows == _EXPECTED_ROLE_COUNTS[role]
        role_total += data_rows
    assert role_total == _EXPECTED_ISSUE_COUNT

    team_rows = _read_csv(tmp_path / "team.csv")
    assert len(team_rows) - 1 == _EXPECTED_ISSUE_COUNT


def test_summary_csv_reports_totals(tmp_path: Path) -> None:
    generate_sheets(_findings(), _meta(), tmp_path)
    rows = _read_csv(tmp_path / SUMMARY_CSV)
    flat = {(r[0], r[1]): r[2] for r in rows if len(r) == 3}
    assert flat[("Totals", "Findings analyzed")] == "13"
    assert flat[("Totals", "Remediation issues")] == str(_EXPECTED_ISSUE_COUNT)
    assert flat[("By owner role", "Technical / Developer")] == str(_EXPECTED_ROLE_COUNTS[ROLE_DEV])


# --- degrade-safe -------------------------------------------------------------
def test_load_findings_is_degrade_safe(tmp_path: Path) -> None:
    assert load_findings(None) is None
    assert load_findings(tmp_path / "missing.json") is None
    bad = tmp_path / "bad.json"
    bad.write_text("{not json", encoding="utf-8")
    assert load_findings(bad) is None
    obj = tmp_path / "obj.json"
    obj.write_text(json.dumps({"not": "a list"}), encoding="utf-8")
    assert load_findings(obj) is None


def test_store_audit_sheets_skips_when_findings_absent(tmp_path: Path) -> None:
    store = LocalArtifactStore(tmp_path / "root")
    # Missing findings => no files, empty return (never raises, never fails a run).
    assert store_audit_sheets(store, "aud-x", None, _meta()) == []
    assert not (tmp_path / "root" / "aud-x" / "sheets").exists()
    # Present findings => the full set under <root>/<audit_id>/sheets/.
    written = store_audit_sheets(store, "aud-1", _FIXTURE, _meta())
    assert set(written) == set(SHEET_FILES)
    assert store.resolve_sheet("aud-1", XLSX_NAME) is not None


# --- staff download route -----------------------------------------------------
class _FakeAuditsRepo:
    def __init__(self, row: dict[str, Any] | None) -> None:
        self.row = row

    def get_audit(self, audit_id: str) -> dict[str, Any] | None:
        return self.row


def _staff_user() -> CurrentUser:
    return CurrentUser(
        id="u-1", email="op@x.com", role="viewer", status="active",
        name="Op", title="", avatar_color="#000", phone="", two_fa=False,
    )


def _seed_store(tmp_path: Path, audit_id: str) -> LocalArtifactStore:
    store = LocalArtifactStore(tmp_path / "root")
    store_audit_sheets(store, audit_id, _FIXTURE, _meta())
    return store


@pytest.fixture
def wire_staff(app: FastAPI) -> Callable[[dict[str, Any] | None, LocalArtifactStore | None], None]:
    def _wire(row: dict[str, Any] | None, store: LocalArtifactStore | None) -> None:
        app.dependency_overrides[get_current_user] = _staff_user
        app.dependency_overrides[get_audits_repo] = lambda: _FakeAuditsRepo(row)
        app.dependency_overrides[get_artifact_store] = lambda: store

    return _wire


async def test_staff_downloads_xlsx_and_csv(
    client: httpx.AsyncClient, tmp_path: Path,
    wire_staff: Callable[[dict[str, Any] | None, LocalArtifactStore | None], None],
) -> None:
    store = _seed_store(tmp_path, "aud-1")
    wire_staff({"id": "aud-1"}, store)

    xlsx = await client.get("/api/v1/audits/aud-1/sheets/remediation.xlsx")
    assert xlsx.status_code == 200
    assert xlsx.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert xlsx.content[:2] == b"PK"  # a zip/xlsx magic

    csv_resp = await client.get("/api/v1/audits/aud-1/sheets/reference.csv")
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    assert "ISS-001" in csv_resp.text


async def test_staff_sheet_rejects_unknown_name_and_missing(
    client: httpx.AsyncClient, tmp_path: Path,
    wire_staff: Callable[[dict[str, Any] | None, LocalArtifactStore | None], None],
) -> None:
    store = _seed_store(tmp_path, "aud-1")
    wire_staff({"id": "aud-1"}, store)
    # Not in the allow-list => 404 (also blocks any traversal attempt).
    assert (await client.get("/api/v1/audits/aud-1/sheets/secrets.csv")).status_code == 404
    assert (await client.get("/api/v1/audits/aud-1/sheets/..%2f..%2fx")).status_code == 404
    # Unknown audit => 404, sheet never resolved.
    wire_staff(None, store)
    assert (await client.get("/api/v1/audits/nope/sheets/team.csv")).status_code == 404


async def test_staff_sheet_requires_view_reports(
    app: FastAPI, client: httpx.AsyncClient, tmp_path: Path,
    wire_staff: Callable[[dict[str, Any] | None, LocalArtifactStore | None], None],
) -> None:
    store = _seed_store(tmp_path, "aud-1")
    wire_staff({"id": "aud-1"}, store)

    def _client_user() -> CurrentUser:
        return CurrentUser(
            id="c-1", email="c@x.com", role="client", status="active",
            name="C", title="", avatar_color="#000", phone="", two_fa=False,
            client_id="cl-A",
        )

    app.dependency_overrides[get_current_user] = _client_user
    # A portal client holds no staff perm => confined out of the staff namespace.
    assert (await client.get("/api/v1/audits/aud-1/sheets/remediation.xlsx")).status_code == 403


# --- portal download route ----------------------------------------------------
class _FakePortalRepo:
    def __init__(self, owned_ids: set[str]) -> None:
        self.owned_ids = owned_ids

    def get_audit(self, audit_id: str) -> dict[str, Any] | None:
        return {"id": audit_id} if audit_id in self.owned_ids else None


def _portal_client() -> CurrentUser:
    return CurrentUser(
        id="u-1", email="p@acme.com", role="client", status="active",
        name="Acme Portal", title="", avatar_color="#000", phone="", two_fa=False,
        client_id="cl-A",
    )


async def test_portal_download_verifies_ownership_then_serves(
    app: FastAPI, client: httpx.AsyncClient, tmp_path: Path
) -> None:
    store = _seed_store(tmp_path, "aud-own")
    app.dependency_overrides[get_current_user] = lambda: _portal_client()
    app.dependency_overrides[get_portal_repo] = lambda: _FakePortalRepo({"aud-own"})
    app.dependency_overrides[get_artifact_store] = lambda: store

    ok = await client.get("/api/v1/portal/audits/aud-own/sheets/remediation.xlsx")
    assert ok.status_code == 200
    assert ok.content[:2] == b"PK"

    # A foreign/unknown id the view does not return => 404, path never resolved.
    missing = await client.get("/api/v1/portal/audits/aud-foreign/sheets/remediation.xlsx")
    assert missing.status_code == 404
    # An unknown sheet name => 404.
    bad = await client.get("/api/v1/portal/audits/aud-own/sheets/secrets.csv")
    assert bad.status_code == 404


async def test_portal_sheet_forbidden_for_staff(
    app: FastAPI, client: httpx.AsyncClient, tmp_path: Path
) -> None:
    store = _seed_store(tmp_path, "aud-own")
    app.dependency_overrides[get_artifact_store] = lambda: store

    def _staff() -> CurrentUser:
        return CurrentUser(
            id="s-1", email="s@x.com", role="owner", status="active",
            name="Staff", title="", avatar_color="#000", phone="", two_fa=False,
        )

    app.dependency_overrides[get_current_user] = _staff
    # Even owner is not a client => 403 out of the /portal namespace.
    resp = await client.get("/api/v1/portal/audits/aud-own/sheets/remediation.xlsx")
    assert resp.status_code == 403
