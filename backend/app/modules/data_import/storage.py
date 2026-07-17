"""Import upload storage: a traversal-safe controlled root for uploaded files.

Copies ``app/services/audit_artifacts.py``'s pattern rather than inventing one - a
generated key under a root, ``resolve`` refusing anything that escapes it - and adds the
two things an UPLOAD needs that an artifact copy does not:

* the name is GENERATED, never derived from the upload. ``audit_artifacts`` keys off an
  internal ``audit_id``; here the only name on offer is attacker-controlled, so it is
  not used at all. ``new_key`` mints ``<uuid4hex>.<ext>`` from a validated extension, so
  ``"../../etc/passwd"``, ``"C:\\Windows\\x"`` and ``"a/../../b"`` are simply never the
  name of anything. The original name survives ONLY as ``import_runs.filename``, a
  display string.
* the write is STREAMED and byte-capped. The size cap is enforced on the way in (a
  ``Content-Length`` can lie, and reading a 10GB body to find out is the attack), so an
  oversized upload is refused mid-stream with the partial file removed.

The ``ImportStore`` seam mirrors ``ArtifactStore`` so an object-store backend can slot
in later without touching the router or the worker.
"""

from __future__ import annotations

import csv
import hashlib
import re
import uuid
import zipfile
from collections.abc import Iterator
from pathlib import Path
from typing import Any, Protocol

from app.config import Settings
from app.modules.data_import.constants import ALLOWED_EXTENSIONS

# 64KB: big enough that a 25MB upload is ~400 reads, small enough that the cap is
# enforced long before a hostile body is resident in memory.
CHUNK_BYTES = 64 * 1024

# csv fields default to a 128KB cap; a pathological cell should be a row error, not a
# crash. Kept modest on purpose - a legitimate export cell is never this big.
MAX_FIELD_BYTES = 1024 * 1024

# A generated key and nothing else: 32 hex chars, a dot, a lowercase extension. The
# store re-checks every key against this before touching the filesystem, so even a
# caller that fabricates a key cannot express a separator, a dot-segment or a drive.
_KEY_RE = re.compile(r"^[0-9a-f]{32}\.[a-z0-9]{1,8}$")


class ImportTooLargeError(RuntimeError):
    """The upload exceeded the byte cap. Raised mid-stream, after the partial file has
    been removed; the router turns it into a 413."""


class ImportRejectedError(RuntimeError):
    """The upload is not storable (a bad extension, or a key that escapes the root)."""


class ImportUnreadableError(RuntimeError):
    """The stored file cannot be parsed as the format its (server-minted) name claims.

    Raised in preference to letting a parser's own exception escape, because those carry
    a traceback that holds the parser's INTERNAL file handle alive - and on Windows an
    open handle makes the cleanup ``unlink`` fail silently, orphaning the upload.
    """


class UploadSource(Protocol):
    """The async chunked reader an upload arrives on.

    Typed structurally rather than as ``fastapi.UploadFile`` so the streaming write is
    unit-testable with a plain fake - and so this module never imports the web layer.
    """

    async def read(self, size: int = -1) -> bytes: ...


class ImportStore(Protocol):
    """Persist an uploaded file under a controlled root; resolve a key back to a path."""

    def new_key(self, extension: str) -> str: ...
    def resolve(self, key: str) -> Path | None: ...


class LocalImportStore:
    """Stores uploads as ``<root>/<uuid4hex>.<ext>`` on a shared filesystem.

    Targets the single-VPS deploy where the API + worker share a disk (like
    ``LocalArtifactStore``): the API writes the upload, the worker streams it back.
    """

    def __init__(self, root: str | Path) -> None:
        self._root = Path(root)

    def new_key(self, extension: str) -> str:
        """Mint a fresh storage key for a validated extension.

        The extension must be in the module's allow-list; the name itself is random, so
        no part of the uploader's input reaches the filesystem.
        """
        ext = extension.lower().lstrip(".")
        if ext not in ALLOWED_EXTENSIONS:
            raise ImportRejectedError(f"Unsupported file type '.{ext}'")
        return f"{uuid.uuid4().hex}.{ext}"

    def path_for(self, key: str) -> Path:
        """The absolute path a key writes to, or raise :class:`ImportRejectedError`.

        Two independent guards, and both are wanted: the ``_KEY_RE`` shape check (a key
        cannot even SPELL a separator or a ``..``), and the resolved-prefix check that
        ``audit_artifacts.resolve`` uses (which catches anything the shape check did not
        anticipate - a symlinked root, a Windows 8.3 alias, a future key format).
        """
        if not _KEY_RE.match(key or ""):
            raise ImportRejectedError("Invalid storage key")
        root = self._root.resolve()
        target = (self._root / key).resolve()
        if not target.is_relative_to(root):
            raise ImportRejectedError("Storage key escapes the import root")
        return target

    def resolve(self, key: str) -> Path | None:
        """Resolve a stored key to a real file within the root, or ``None``.

        Refuses any key that escapes the root (``..`` / absolute / a separator), so a
        crafted ``stored_path`` can never make the worker read an arbitrary file.
        Mirrors ``LocalArtifactStore.resolve`` exactly.
        """
        try:
            target = self.path_for(key)
        except ImportRejectedError:
            return None
        return target if target.is_file() else None

    def delete(self, key: str) -> None:
        """Best-effort removal of a stored upload (a rejected/oversized write). Never
        raises: cleanup failing must not mask the reason we are cleaning up."""
        try:
            path = self.path_for(key)
            path.unlink(missing_ok=True)
        except (ImportRejectedError, OSError):
            return


async def write_upload(
    store: LocalImportStore, key: str, upload: UploadSource, *, max_bytes: int
) -> tuple[int, str]:
    """Stream ``upload`` into ``key`` under the store's root; return ``(bytes, sha256)``.

    The cap is enforced as the bytes ARRIVE, not from a header: a ``Content-Length`` is
    a claim by the client, and a chunked body has none at all. Crossing the cap removes
    the partial file and raises :class:`ImportTooLargeError`, so a hostile upload costs the
    cap plus one chunk of disk and nothing more. Any other failure cleans up too, so a
    half-written file is never left behind for the worker to parse.
    """
    path = store.path_for(key)
    path.parent.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256()
    total = 0
    try:
        with path.open("wb") as fh:
            while True:
                chunk = await upload.read(CHUNK_BYTES)
                if not chunk:
                    break
                total += len(chunk)
                if total > max_bytes:
                    raise ImportTooLargeError(f"Upload exceeds the {max_bytes} byte limit")
                digest.update(chunk)
                fh.write(chunk)
    except BaseException:
        store.delete(key)
        raise
    return total, digest.hexdigest()


# --------------------------------------------------------------------------- #
# Streaming readers.
# --------------------------------------------------------------------------- #
# These live HERE, beside the store that owns the file, rather than in ``tasks.py``.
# The worker is not the only reader: the upload route previews a file's header + sample
# values the moment it lands. Importing them from ``tasks`` would drag ``celery_app``
# into the API process on every upload - exactly what the module's lazily-imported
# enqueuer exists to avoid. ``service.py`` cannot host them either: it is pure (no
# filesystem), and that is worth keeping.
def iter_csv(path: Path, *, delimiter: str) -> Iterator[list[Any]]:
    """Stream a CSV/TSV row by row.

    ``newline=""`` is required by the csv module (it does its own line handling);
    ``utf-8-sig`` transparently eats the BOM Excel writes, which would otherwise turn the
    first header into ``"﻿Query"`` and silently break auto-mapping.
    ``errors="replace"`` keeps a stray legacy byte from killing an otherwise fine file.
    """
    with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as fh:
        yield from csv.reader(fh, delimiter=delimiter)


def iter_xlsx(path: Path) -> Iterator[list[Any]]:
    """Stream an XLSX row by row.

    ``read_only=True`` is what makes this a stream: openpyxl parses the sheet lazily
    instead of building the whole workbook in memory. ``data_only=True`` reads a formula
    cell's cached VALUE rather than its formula text.

    The container is validated FIRST, in a ``with`` block that closes deterministically.
    This is not belt-and-braces - it closes a real gap. An xlsx IS a zip, so a plain zip
    (a folder of photos) renamed ``.xlsx`` passes the upload sniff, and openpyxl then
    raises ``KeyError: '[Content_Types].xml'`` from deep inside itself. That traceback
    holds openpyxl's own ZipFile alive, and on Windows an open handle makes the caller's
    cleanup ``unlink`` fail - so the rejected upload is orphaned in the import root.
    Checking here means the common case raises a clean, shallow error instead.
    """
    from openpyxl import load_workbook

    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
    except (zipfile.BadZipFile, OSError) as exc:
        raise ImportUnreadableError("the file is not a readable .xlsx container") from exc
    if "[Content_Types].xml" not in names:
        raise ImportUnreadableError("the file is a zip archive, but not an .xlsx workbook")

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ImportUnreadableError("the workbook could not be opened") from exc
    try:
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            yield list(row)
    finally:
        wb.close()


def iter_rows(path: Path) -> Iterator[list[Any]]:
    """Stream a stored file's rows, dispatching on its stored extension.

    The extension is trustworthy HERE (unlike at upload): the store minted this name
    itself, after the route sniffed the bytes - so it is the server's own record of what
    the file turned out to be, not the uploader's claim.
    """
    csv.field_size_limit(MAX_FIELD_BYTES)
    suffix = path.suffix.lower().lstrip(".")
    if suffix == "xlsx":
        return iter_xlsx(path)
    return iter_csv(path, delimiter="\t" if suffix == "tsv" else ",")


def read_head(path: Path, size: int = 8192) -> bytes:
    """The first bytes of a stored upload - the sniffer's input. Blocking; offload it."""
    with path.open("rb") as fh:
        head: bytes = fh.read(size)
    return head


def import_store_from_settings(settings: Settings) -> LocalImportStore | None:
    """Build the local import store, or ``None`` when unconfigured.

    Mirrors ``audit_artifacts.local_store_from_settings``: an unconfigured root is a
    DEGRADED module (the upload route reports "not configured"), never a crash and never
    a silent write to some default directory.
    """
    root = settings.import_artifact_dir
    return LocalImportStore(root) if root else None
