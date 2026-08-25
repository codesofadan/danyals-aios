"""The keyword workbook - one tab per page, and a method trail that survives (P7).

THE REQUIREMENT IN THE OWNER'S OWN WORDS: keyword research exported as a spreadsheet,
one tab per page, documenting keywords, clusters, strategy and method "so that we can
audit it months later". That last clause is the hard part, and it is what the METHOD &
SOURCES tab exists for.

WHAT MAKES IT AUDITABLE RATHER THAN MERELY DETAILED. Every number carries where it came
from. `keyword_plan_terms.estimated` is a real column, so a figure we DERIVED is
labelled as derived and a figure we BOUGHT is labelled with the provider and the date
it was pulled. A spreadsheet that presents both as the same kind of fact is the thing
this whole rebuild exists to stop: v1 computed `difficulty = log10(totalResults) * 8`
and shipped it looking exactly like vendor data.

The doctrine trail comes from `doctrine_usage`, which records - per stage, per page -
which chunks governed the writing, which model ran, what it cost, and what was DROPPED
when a pack would not fit. Months later that answers "why does this page say this",
which no amount of prose in a README can.

NO NEW DEPENDENCY. `openpyxl>=3.1` is already core (`pyproject.toml:36`).

TWO OPENPYXL TRAPS, BOTH HANDLED IN `sheet_title`:
  - a sheet title is capped at 31 characters and Excel silently refuses longer ones
  - []:*?/\\ are illegal in a title, and a keyword like "roof repair / gutters" has one

A page tab is named from its keyword, so both traps are live on real data rather than
theoretical. Uniqueness is enforced with a numeric prefix, because two pages in one
engagement genuinely can share a truncated name.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

# Excel's own limits, not ours.
MAX_SHEET_TITLE = 31
_ILLEGAL_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")

README_HEADERS = ("Field", "Value")
METHOD_HEADERS = ("Step", "Source", "Detail", "Count", "Cost (USD)", "When")
MASTER_HEADERS = (
    "Keyword", "Volume", "Difficulty", "CPC", "Competition", "Intent",
    "Relevance", "Opportunity", "Cluster", "Data source",
)
CLUSTER_HEADERS = ("Cluster", "Terms", "Total volume", "Top term", "Intent")
MAP_HEADERS = (
    "Silo", "Page type", "Primary keyword", "Intent", "City", "Priority",
    "Target words", "Cluster", "Evidence", "Information-gain thesis", "Status",
)
PAGE_HEADERS = ("Field", "Value")
DOCTRINE_HEADERS = (
    "Page", "Stage", "Model", "Chunks used", "Chunks dropped",
    "Input tokens", "Output tokens", "Cache read", "Cost (USD)", "When",
)

# The label that appears in the "Data source" column. Deliberately not a bare boolean:
# a reader scanning the column must not have to know what `estimated = TRUE` implies.
VENDOR_LABEL = "measured - {provider}"
ESTIMATED_LABEL = "ESTIMATED - derived, not vendor data"


def sheet_title(name: str, used: set[str]) -> str:
    """A legal, unique Excel sheet title for ``name``.

    Strips the characters Excel forbids, trims to 31, and disambiguates collisions with
    a numeric prefix rather than a suffix - a suffix would be the first thing cut off
    by the very truncation that caused the collision.
    """
    cleaned = _ILLEGAL_TITLE_CHARS.sub("-", (name or "").strip()) or "Sheet"
    cleaned = re.sub(r"\s+", " ", cleaned).strip("'")
    candidate = cleaned[:MAX_SHEET_TITLE]
    if candidate.lower() not in used:
        used.add(candidate.lower())
        return candidate
    for n in range(2, 1000):
        prefix = f"{n} "
        candidate = f"{prefix}{cleaned[: MAX_SHEET_TITLE - len(prefix)]}"
        if candidate.lower() not in used:
            used.add(candidate.lower())
            return candidate
    raise ValueError(f"could not make a unique sheet title for {name!r}")


def data_source(term: dict[str, Any], provider: str) -> str:
    """How this row's numbers were obtained, in words a reader can act on."""
    if term.get("estimated"):
        return ESTIMATED_LABEL
    return VENDOR_LABEL.format(provider=provider or "provider not recorded")


def master_rows(terms: list[dict[str, Any]], provider: str) -> list[list[Any]]:
    return [
        [
            t.get("keyword", ""), t.get("volume"), t.get("difficulty"), t.get("cpc"),
            t.get("competition"), t.get("intent", ""), t.get("relevance"),
            t.get("opportunity"), t.get("cluster_key", ""), data_source(t, provider),
        ]
        for t in terms
    ]


def cluster_rows(terms: list[dict[str, Any]]) -> list[list[Any]]:
    """One row per cluster. Volume sums SKIP estimated terms deliberately.

    Adding a derived number to a bought one produces a total that is neither, and it is
    the totals people quote in meetings.
    """
    buckets: dict[str, list[dict[str, Any]]] = {}
    for t in terms:
        buckets.setdefault(t.get("cluster_key") or "(unclustered)", []).append(t)

    rows: list[list[Any]] = []
    for key, group in sorted(buckets.items()):
        measured = [t for t in group if not t.get("estimated") and t.get("volume")]
        total = sum(int(t["volume"]) for t in measured)
        top = max(group, key=lambda t: (t.get("volume") or 0))
        intents = sorted({t.get("intent", "") for t in group if t.get("intent")})
        rows.append([
            key, len(group),
            total if measured else "n/a - all estimated",
            top.get("keyword", ""), ", ".join(intents),
        ])
    return rows


def map_rows(nodes: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [
            n.get("silo", ""), n.get("page_type", ""), n.get("primary_keyword", ""),
            n.get("intent", ""), n.get("target_city", ""), n.get("priority", 0),
            n.get("target_words", 0), n.get("cluster_key", ""),
            n.get("evidence", ""), n.get("info_gain_thesis", ""),
            "published" if n.get("published_url") else
            ("in production" if n.get("content_job_id") else "planned"),
        ]
        for n in nodes
    ]


def page_rows(
    node: dict[str, Any], terms: list[dict[str, Any]], provider: str
) -> list[list[Any]]:
    """One page's tab: the plan for it, and the terms it is responsible for.

    Vertical (field/value) rather than a wide table, because this tab is read by a
    human deciding whether the page is right - not scanned or sorted.
    """
    secondary = node.get("secondary_keywords") or []
    rows: list[list[Any]] = [
        ["Primary keyword", node.get("primary_keyword", "")],
        ["Page type", node.get("page_type", "")],
        ["Silo", node.get("silo", "")],
        ["Search intent", node.get("intent", "")],
        ["Target city", node.get("target_city", "")],
        ["Target words", node.get("target_words", 0)],
        ["Priority", node.get("priority", 0)],
        ["Cluster", node.get("cluster_key", "")],
        ["Why this page can rank", node.get("evidence", "")],
        ["What it adds that the top 10 does not", node.get("info_gain_thesis", "")],
        ["Published URL", node.get("published_url", "") or "not yet published"],
        [],
        ["Secondary keywords", f"{len(secondary)} term(s)"],
    ]
    rows.extend([["", s] for s in secondary])
    rows.append([])
    rows.append(["Assigned terms", f"{len(terms)} term(s)"])
    rows.append(list(MASTER_HEADERS))
    rows.extend(master_rows(terms, provider))
    return rows


def doctrine_rows(usage: list[dict[str, Any]], page_by_job: dict[str, str]) -> list[list[Any]]:
    """The method trail: what governed each page, per stage.

    Dropped chunks are shown, not hidden. A pack that did not fit is a real limit on
    what the model could see, and a trail that omits it overstates the doctrine's reach.
    """
    return [
        [
            page_by_job.get(str(u.get("job_id")), "(engagement-level)"),
            u.get("stage", ""), u.get("model", ""),
            len(u.get("chunk_ids") or []), len(u.get("dropped_chunk_ids") or []),
            u.get("input_tokens", 0), u.get("output_tokens", 0),
            u.get("cache_read_tokens", 0), u.get("cost", 0), _ts(u.get("created_at")),
        ]
        for u in usage
    ]


def method_rows(
    plan: dict[str, Any], terms: list[dict[str, Any]], usage: list[dict[str, Any]]
) -> list[list[Any]]:
    """The tab that makes the rest auditable: what was bought, what was derived."""
    measured = [t for t in terms if not t.get("estimated")]
    derived = [t for t in terms if t.get("estimated")]
    provider = plan.get("provider") or "not recorded"
    pulled = _ts(plan.get("provider_run_at"))

    rows: list[list[Any]] = [
        ["Keyword metrics (bought)", provider,
         "volume, difficulty, CPC and competition as returned by the provider",
         len(measured), plan.get("cost", 0), pulled],
    ]
    if derived:
        rows.append([
            "Keyword metrics (derived)", "none - computed locally",
            "no provider row for these terms; volume and difficulty are ESTIMATES and "
            "must not be quoted as measured demand",
            len(derived), 0, pulled,
        ])
    rows.append([
        "Seed terms", "operator brief",
        ", ".join(plan.get("seed_terms") or []) or "(none recorded)",
        len(plan.get("seed_terms") or []), 0, _ts(plan.get("created_at")),
    ])

    by_stage: dict[str, dict[str, Any]] = {}
    for u in usage:
        agg = by_stage.setdefault(
            u.get("stage", ""), {"calls": 0, "cost": 0.0, "model": u.get("model", "")}
        )
        agg["calls"] += 1
        agg["cost"] += float(u.get("cost") or 0)
    for stage, agg in sorted(by_stage.items()):
        rows.append([
            f"Writing stage: {stage}", agg["model"] or "model not recorded",
            "doctrine-governed generation; see the Method Trail tab for the chunks",
            agg["calls"], round(agg["cost"], 4), "",
        ])

    total = float(plan.get("cost") or 0) + sum(a["cost"] for a in by_stage.values())
    rows.append([])
    rows.append(["TOTAL", "", "every metered call recorded for this engagement", "",
                 round(total, 4), ""])
    return rows


def readme_rows(
    engagement: dict[str, Any], terms: list[dict[str, Any]], nodes: list[dict[str, Any]]
) -> list[list[Any]]:
    derived = sum(1 for t in terms if t.get("estimated"))
    rows = [
        ["Client", engagement.get("client_name", "")],
        ["Engagement", engagement.get("shape", "")],
        ["Generated", _ts(engagement.get("generated_at"))],
        [],
        ["Keywords", len(terms)],
        ["  of which measured", len(terms) - derived],
        ["  of which ESTIMATED", derived],
        ["Planned pages", len(nodes)],
        [],
        ["How to read this workbook", ""],
        ["Master Keywords",
         "every term, with a Data source column saying whether the numbers were bought "
         "or derived"],
        ["Method & Sources",
         "what was paid for, from whom, when, and what it cost"],
        ["Method Trail",
         "which doctrine governed which page, per writing stage"],
        ["One tab per page",
         "the plan for that page and the terms it is responsible for"],
        [],
        ["A note on ESTIMATED rows",
         "search volume originates in Google's ad auction and cannot be derived "
         "offline. A row marked ESTIMATED is our own approximation and must not be "
         "quoted to a client as measured demand."],
    ]
    return rows


def _ts(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return text[:19] if len(text) >= 19 else text


@dataclass
class WorkbookResult:
    path: Path | None = None
    sheets: list[str] = field(default_factory=list)
    terms: int = 0
    pages: int = 0
    estimated_terms: int = 0


def render(
    data: dict[str, Any],
    out_path: str | Path,
) -> WorkbookResult:
    """Write the workbook. Pure: everything it needs is in ``data``.

    Separated from the database fetch on purpose - the tab layout, the sheet-title
    rules and the measured/estimated labelling are the parts worth testing, and none of
    them need a connection.
    """
    from openpyxl import Workbook

    engagement = data.get("engagement") or {}
    plan = data.get("plan") or {}
    terms: list[dict[str, Any]] = data.get("terms") or []
    nodes: list[dict[str, Any]] = data.get("nodes") or []
    usage: list[dict[str, Any]] = data.get("usage") or []
    provider = plan.get("provider") or ""

    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    result = WorkbookResult(
        path=path, terms=len(terms), pages=len(nodes),
        estimated_terms=sum(1 for t in terms if t.get("estimated")),
    )

    def add(title: str, headers: tuple[str, ...] | None, rows: list[list[Any]]) -> None:
        ws = wb.create_sheet(sheet_title(title, used))
        result.sheets.append(ws.title)
        if headers:
            ws.append(list(headers))
        for row in rows:
            ws.append(row)

    add("README", README_HEADERS, readme_rows(engagement, terms, nodes))
    add("Method & Sources", METHOD_HEADERS, method_rows(plan, terms, usage))
    add("Master Keywords", MASTER_HEADERS, master_rows(terms, provider))
    add("Clusters", CLUSTER_HEADERS, cluster_rows(terms))
    add("Topical Map", MAP_HEADERS, map_rows(nodes))

    terms_by_cluster: dict[str, list[dict[str, Any]]] = {}
    for t in terms:
        terms_by_cluster.setdefault(t.get("cluster_key") or "", []).append(t)

    page_by_job: dict[str, str] = {}
    for node in nodes:
        title = node.get("primary_keyword") or node.get("page_type") or "page"
        assigned = terms_by_cluster.get(node.get("cluster_key") or "", [])
        ws_title = sheet_title(title, used)
        ws = wb.create_sheet(ws_title)
        result.sheets.append(ws_title)
        ws.append(list(PAGE_HEADERS))
        for row in page_rows(node, assigned, provider):
            ws.append(row)
        if node.get("content_job_id"):
            page_by_job[str(node["content_job_id"])] = ws_title

    add("Method Trail", DOCTRINE_HEADERS, doctrine_rows(usage, page_by_job))
    wb.save(path)
    return result
